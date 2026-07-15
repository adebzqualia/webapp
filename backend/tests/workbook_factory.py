from __future__ import annotations

import copy
import io
from typing import Callable

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo


def workbook_bytes(single_sheet: bool = False) -> bytes:
    workbook = Workbook()
    financial = workbook.active
    financial.title = "Financial KPIs"
    financial.merge_cells("A1:D1")
    financial["A1"] = "Financial performance"
    financial["A1"].font = Font(bold=True, size=14)
    financial["B3"] = "Metric"
    financial["C3"] = "Budget"
    financial["D3"] = "Actual"
    for cell in financial[3][1:4]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    rows = [
        ("Sales", 100, 110),
        ("Opex", 40, 45),
        ("Profit", 60, "=D4-D5"),
    ]
    for row_number, values in enumerate(rows, start=4):
        for column, value in enumerate(values, start=2):
            financial.cell(row_number, column, value)
    revenue_table = Table(displayName="RevenueTable", ref="B3:D6")
    revenue_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
    )
    financial.add_table(revenue_table)

    financial["F3"] = "KPI"
    financial["G3"] = "Value"
    financial["F4"] = "Headcount"
    financial["G4"] = 12
    financial["F5"] = "Sites"
    financial["G5"] = 3

    if not single_sheet:
        operations = workbook.create_sheet("Operations")
        operations.append(["Operations overview"])
        operations.append(["Site", "Units", "Quality"])
        operations.append(["North", 20, 0.98])
        operations.append(["South", 30, 0.96])
        operations.append(["West", 25, 0.97])
        hidden = workbook.create_sheet("Instructions")
        hidden["A1"] = "Do not edit"
        hidden.sheet_state = "hidden"
        workbook.defined_names.add(
            DefinedName("OpsRange", attr_text="'Operations'!$A$2:$C$5")
        )

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def mutate(data: bytes, callback: Callable) -> bytes:
    workbook = load_workbook(io.BytesIO(data), data_only=False)
    callback(workbook)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _copy_rectangle(ws, source_ref: str, target_top_left: str, *, clear_source: bool = True):
    min_col, min_row, max_col, max_row = range_boundaries(source_ref)
    target = ws[target_top_left]
    for row_offset, row in enumerate(range(min_row, max_row + 1)):
        for col_offset, col in enumerate(range(min_col, max_col + 1)):
            source_cell = ws.cell(row, col)
            target_cell = ws.cell(target.row + row_offset, target.column + col_offset)
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell._style = copy.copy(source_cell._style)
            if clear_source:
                source_cell.value = None


def moved_table(data: bytes) -> bytes:
    def change(workbook):
        ws = workbook["Financial KPIs"]
        _copy_rectangle(ws, "B3:D6", "H10")
        ws.tables["RevenueTable"].ref = "H10:J13"

    return mutate(data, change)


def ambiguous_tables(data: bytes) -> bytes:
    def change(workbook):
        ws = workbook["Financial KPIs"]
        _copy_rectangle(ws, "B3:D6", "H10")
        _copy_rectangle(ws, "H10:J13", "N10", clear_source=False)
        del ws.tables["RevenueTable"]

    return mutate(data, change)


def renamed_sheet(data: bytes) -> bytes:
    return mutate(data, lambda workbook: setattr(workbook["Financial KPIs"], "title", "Finance Country"))


def removed_sheet(data: bytes) -> bytes:
    return mutate(data, lambda workbook: workbook.remove(workbook["Operations"]))


def added_sheet(data: bytes) -> bytes:
    def change(workbook):
        workbook.create_sheet("Country Notes")["A1"] = "Extra"

    return mutate(data, change)


def reordered_sheets(data: bytes) -> bytes:
    def change(workbook):
        workbook._sheets.insert(0, workbook._sheets.pop(1))

    return mutate(data, change)


def added_column(data: bytes) -> bytes:
    def change(workbook):
        ws = workbook["Financial KPIs"]
        ws["E3"] = "Forecast"
        ws["E4"] = 105
        ws["E5"] = 43
        ws["E6"] = 62
        ws.tables["RevenueTable"].ref = "B3:E6"

    return mutate(data, change)


def removed_column(data: bytes) -> bytes:
    def change(workbook):
        ws = workbook["Financial KPIs"]
        ws.delete_cols(3, 1)
        ws.tables["RevenueTable"].ref = "B3:C6"

    return mutate(data, change)


def reversed_columns(data: bytes) -> bytes:
    def change(workbook):
        ws = workbook["Financial KPIs"]
        for row in range(3, 7):
            ws.cell(row, 2).value, ws.cell(row, 3).value = (
                ws.cell(row, 3).value,
                ws.cell(row, 2).value,
            )

    return mutate(data, change)


def added_row(data: bytes) -> bytes:
    def change(workbook):
        ws = workbook["Financial KPIs"]
        ws.append([])
        ws["B7"] = "Tax"
        ws["C7"] = 10
        ws["D7"] = 9
        ws.tables["RevenueTable"].ref = "B3:D7"

    return mutate(data, change)


def removed_row(data: bytes) -> bytes:
    def change(workbook):
        ws = workbook["Financial KPIs"]
        ws.delete_rows(5, 1)
        ws.tables["RevenueTable"].ref = "B3:D5"

    return mutate(data, change)


def reversed_rows(data: bytes) -> bytes:
    def change(workbook):
        ws = workbook["Financial KPIs"]
        for column in range(2, 5):
            ws.cell(4, column).value, ws.cell(5, column).value = (
                ws.cell(5, column).value,
                ws.cell(4, column).value,
            )

    return mutate(data, change)


def removed_formula(data: bytes) -> bytes:
    def change(workbook):
        workbook["Financial KPIs"]["D6"] = 65

    return mutate(data, change)


def styled_consolidation_workbook(sheet_name: str = "Very long operational performance summary") -> bytes:
    workbook = Workbook()
    ws = workbook.active
    ws.title = sheet_name[:31]
    ws["A1"] = "Header"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = 10
    ws["B2"] = "=A2*2"
    ws.merge_cells("A3:B3")
    ws["A3"] = "Merged"
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 24
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()
