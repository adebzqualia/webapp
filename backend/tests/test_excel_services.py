from __future__ import annotations

import io
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.config import Settings
from app.errors import DomainError
from app.services.consolidator import (
    ConsolidationSource,
    OpenpyxlWorkbookConsolidator,
    safe_sheet_name,
)
from app.services.excel_reader import OpenpyxlWorkbookReader
from app.services.file_security import XlsxSecurityValidator
from app.services.table_detector import HeuristicTableDetector

from .workbook_factory import styled_consolidation_workbook, workbook_bytes


def test_inspection_single_and_multiple_sheets(tmp_path: Path):
    reader = OpenpyxlWorkbookReader()
    single = tmp_path / "single.xlsx"
    single.write_bytes(workbook_bytes(single_sheet=True))
    single_metadata = reader.inspect_workbook(single)
    assert single_metadata["sheetCount"] == 1
    assert single_metadata["sheets"][0]["formulaCount"] == 1
    assert single_metadata["sheets"][0]["mergedRanges"] == ["A1:D1"]
    assert single_metadata["sheets"][0]["nativeTables"][0]["rangeRef"] == "B3:D6"

    multiple = tmp_path / "multiple.xlsx"
    multiple.write_bytes(workbook_bytes())
    metadata = reader.inspect_workbook(multiple)
    assert metadata["sheetNames"] == ["Financial KPIs", "Operations", "Instructions"]
    assert metadata["sheets"][2]["visibility"] == "hidden"
    assert any(item["name"] == "OpsRange" for item in metadata["definedNames"])
    assert metadata["formulasEvaluated"] is False


def test_grid_and_structured_preview_keep_formulas_as_text(tmp_path: Path):
    path = tmp_path / "grid.xlsx"
    path.write_bytes(workbook_bytes())
    reader = OpenpyxlWorkbookReader(max_grid_cells=100)
    grid = reader.get_grid(path, "Financial KPIs", "B3:D6")
    formula = next(cell for cell in grid["cells"] if cell["coordinate"] == "D6")
    assert formula["formula"] == "=D4-D5"
    preview = reader.preview_range(path, "Financial KPIs", "B3:D6", [3], 4, 6, ["B"])
    assert preview["headers"] == ["Metric", "Budget", "Actual"]
    assert preview["estimatedTypes"] == ["text", "number", "mixed"]
    assert preview["rows"][2]["Actual"] == "=D4-D5"


def test_detector_finds_multiple_tables_on_same_sheet(tmp_path: Path):
    path = tmp_path / "tables.xlsx"
    path.write_bytes(workbook_bytes())
    candidates = HeuristicTableDetector().detect_candidates(path, ["Financial KPIs"])
    ranges = {candidate["rangeRef"] for candidate in candidates}
    assert "B3:D6" in ranges
    assert "F3:G5" in ranges
    assert all(0 <= candidate["confidence"] <= 1 for candidate in candidates)
    assert all(candidate["reasons"] for candidate in candidates)


def test_file_security_rejects_extension_signature_corruption_and_zip_bomb():
    validator = XlsxSecurityValidator(Settings())
    valid = workbook_bytes(single_sheet=True)
    assert validator.validate("valid.xlsx", valid).sha256
    with pytest.raises(DomainError, match="Only .xlsx"):
        validator.validate("valid.xls", valid)
    with pytest.raises(DomainError) as invalid_signature:
        validator.validate("fake.xlsx", b"not a zip")
    assert invalid_signature.value.code == "INVALID_FILE_SIGNATURE"
    with pytest.raises(DomainError) as corrupted:
        validator.validate("corrupt.xlsx", b"PK\x03\x04broken")
    assert corrupted.value.code == "FILE_CORRUPTED"

    payload = io.BytesIO()
    with zipfile.ZipFile(payload, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", "x")
        archive.writestr("xl/workbook.xml", "x")
        archive.writestr("xl/worksheets/sheet1.xml", "0" * 2_000_000)
    strict = XlsxSecurityValidator(
        Settings(max_zip_compression_ratio=20, max_zip_uncompressed_bytes=3_000_000)
    )
    with pytest.raises(DomainError) as bomb:
        strict.validate("bomb.xlsx", payload.getvalue())
    assert bomb.value.code == "ZIP_BOMB_DETECTED"


def test_sheet_name_sanitization_limits_collisions_and_special_characters():
    used: set[str] = set()
    first = safe_sheet_name("Operational / Performance [Summary]", "Côte d'Ivoire", used)
    second = safe_sheet_name("Operational / Performance [Summary]", "Côte d'Ivoire", used)
    assert len(first) <= 31 and len(second) <= 31
    assert first != second
    assert not set("[]:*?/\\") & set(first)
    assert second.endswith("_2")


def test_consolidation_preserves_core_cells_and_reports_name_mapping(tmp_path: Path):
    source_a = tmp_path / "a.xlsx"
    source_b = tmp_path / "b.xlsx"
    source_a.write_bytes(styled_consolidation_workbook())
    source_b.write_bytes(styled_consolidation_workbook())
    sources = [
        ConsolidationSource("a", "France", "FR", "va", "a.xlsx", source_a),
        ConsolidationSource("b", "France duplicate", "FR", "vb", "b.xlsx", source_b),
    ]
    data, report, filename = OpenpyxlWorkbookConsolidator().consolidate(sources)
    assert filename.endswith(".xlsx")
    assert len(report["nameMappings"]) == 2
    workbook = load_workbook(io.BytesIO(data), data_only=False)
    assert len(workbook.sheetnames) == 2
    assert len(set(workbook.sheetnames)) == 2
    assert all(len(name) <= 31 for name in workbook.sheetnames)
    first = workbook[workbook.sheetnames[0]]
    assert first["B2"].value == "=A2*2"
    assert first["A1"].font.bold is True
    assert "A3:B3" in {str(value) for value in first.merged_cells.ranges}
    assert first.freeze_panes == "A2"
    workbook.close()
