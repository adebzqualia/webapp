from __future__ import annotations

import io

from openpyxl import load_workbook

from .workbook_factory import (
    added_column,
    added_row,
    added_sheet,
    ambiguous_tables,
    moved_table,
    removed_column,
    removed_formula,
    removed_row,
    removed_sheet,
    renamed_sheet,
    reordered_sheets,
    reversed_columns,
    reversed_rows,
    workbook_bytes,
)


def upload_template_and_mapping(client):
    source = workbook_bytes()
    response = client.post(
        "/api/templates",
        data={"name": "POPS Reference"},
        files={"file": ("template.xlsx", source, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 201, response.text
    template = response.json()
    version = template["versions"][-1]
    sheets = version["sheets"]
    financial = next(sheet for sheet in sheets if sheet["name"] == "Financial KPIs")
    table_response = client.post(
        f"/api/templates/{template['id']}/tables",
        json={
            "sheetId": financial["id"],
            "name": "Revenue Budget",
            "rangeRef": "B3:D6",
            "headerRows": [3],
            "dataStartRow": 4,
            "dataEndRow": 6,
            "keyColumns": ["B"],
            "valueColumns": ["C", "D"],
            "requiredFormulas": [{"coordinate": "D6"}],
            "structureMode": "STRICT",
        },
    )
    assert table_response.status_code == 201, table_response.text
    table = table_response.json()
    validated = client.post(
        f"/api/templates/{template['id']}/tables/{table['id']}/validate"
    )
    assert validated.status_code == 200, validated.text
    for sheet in sheets:
        if sheet["id"] != financial["id"]:
            ignored = client.patch(
                f"/api/templates/{template['id']}/sheets/{sheet['id']}",
                json={"ignored": True},
            )
            assert ignored.status_code == 200
    return source, template, financial, table


def create_country(client, template_id: str, name: str, code: str | None = None):
    response = client.post(
        "/api/countries",
        json={"name": name, "code": code, "templateId": template_id},
    )
    assert response.status_code == 201, response.text
    return response.json()


def upload_country(client, country_id: str, data: bytes, filename: str = "country.xlsx"):
    response = client.post(
        f"/api/countries/{country_id}/files",
        data={"autoAnalyze": "false"},
        files={"file": (filename, data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 201, response.text
    return response.json()


def analyze_version(client, version_id: str):
    response = client.post(f"/api/country-files/{version_id}/analyze")
    assert response.status_code == 200, response.text
    assert response.json()["status"] in {"COMPLETED", "FAILED"}
    detail = client.get(f"/api/country-files/{version_id}/analysis")
    assert detail.status_code == 200, detail.text
    return detail.json()


def test_template_import_grid_detection_preview_mapping_and_org_isolation(client):
    _, template, financial, table = upload_template_and_mapping(client)
    listing = client.get("/api/templates").json()
    assert listing[0]["originalFilename"] == "template.xlsx"
    assert listing[0]["status"] == "READY"
    assert listing[0]["tableCount"] == 1
    sheets = client.get(f"/api/templates/{template['id']}/sheets").json()
    assert [sheet["name"] for sheet in sheets] == [
        "Financial KPIs",
        "Operations",
        "Instructions",
    ]
    grid = client.get(
        f"/api/templates/{template['id']}/sheets/{financial['id']}/grid",
        params={"rangeRef": "B3:D6"},
    )
    assert grid.status_code == 200
    assert next(cell for cell in grid.json()["cells"] if cell["coordinate"] == "D6")["formula"] == "=D4-D5"
    detected = client.post(
        f"/api/templates/{template['id']}/tables/detect",
        json={"sheetId": financial["id"]},
    )
    assert {item["rangeRef"] for item in detected.json()} >= {"B3:D6", "F3:G5"}
    preview = client.post(
        f"/api/templates/{template['id']}/tables/preview",
        json={
            "sheetId": financial["id"],
            "rangeRef": "B3:D6",
            "headerRows": [3],
            "dataStartRow": 4,
            "dataEndRow": 6,
            "keyColumns": ["B"],
        },
    )
    assert preview.status_code == 200
    assert preview.json()["headers"] == ["Metric", "Budget", "Actual"]
    export = client.get(f"/api/templates/{template['id']}/mapping").json()
    assert export["workbookHash"]
    assert export["mappingVersion"] >= 5
    assert table["status"] == "DRAFT"
    other_org = client.get(
        f"/api/templates/{template['id']}", headers={"X-Organization-Id": "other-org"}
    )
    assert other_org.status_code == 404


def test_compliant_country_analysis_and_extraction(client):
    source, template, _, _ = upload_template_and_mapping(client)
    country = create_country(client, template["id"], "France", "FR")
    imported = upload_country(client, country["id"], source)
    version_id = imported["currentVersion"]["id"]
    detail = analyze_version(client, version_id)
    assert detail["job"]["status"] == "COMPLETED"
    assert detail["job"]["report"]["blockingCount"] == 0
    assert detail["anomalies"] == []
    assert detail["extractedTables"][0]["rows"][0]["Metric"] == "Sales"
    countries = client.get("/api/countries").json()
    assert countries[0]["currentFile"]["id"] == imported["id"]
    assert countries[0]["status"] == "COMPLIANT"


def test_workbook_and_table_anomaly_categories(client):
    source, template, _, _ = upload_template_and_mapping(client)
    country = create_country(client, template["id"], "Germany", "DE")
    variants = [
        (removed_sheet(source), "SHEET_MISSING"),
        (added_sheet(source), "SHEET_ADDED"),
        (renamed_sheet(source), "SHEET_RENAMED"),
        (reordered_sheets(source), "SHEET_ORDER_CHANGED"),
        (moved_table(source), "TABLE_MOVED"),
        (added_column(source), "COLUMN_ADDED"),
        (removed_column(source), "COLUMN_REMOVED"),
        (reversed_columns(source), "COLUMN_ORDER_CHANGED"),
        (added_row(source), "ROW_ADDED"),
        (removed_row(source), "ROW_REMOVED"),
        (reversed_rows(source), "ROW_ORDER_CHANGED"),
        (removed_formula(source), "FORMULA_MISSING"),
        (ambiguous_tables(source), "AMBIGUOUS_TABLE_MATCH"),
    ]
    observed: dict[str, set[str]] = {}
    for index, (data, expected_category) in enumerate(variants, start=1):
        imported = upload_country(client, country["id"], data, f"variant-{index}.xlsx")
        version_id = imported["currentVersion"]["id"]
        detail = analyze_version(client, version_id)
        categories = {item["category"] for item in detail["anomalies"]}
        observed[expected_category] = categories
    missing = {
        expected: categories
        for expected, categories in observed.items()
        if expected not in categories
    }
    assert not missing, missing


def test_anomaly_decision_dashboard_and_latest_country_filter(client):
    source, template, _, _ = upload_template_and_mapping(client)
    country = create_country(client, template["id"], "Spain", "ES")
    imported = upload_country(client, country["id"], removed_formula(source))
    version_id = imported["currentVersion"]["id"]
    detail = analyze_version(client, version_id)
    anomaly = next(item for item in detail["anomalies"] if item["category"] == "FORMULA_MISSING")
    decision = client.patch(
        f"/api/anomalies/{anomaly['id']}",
        json={"status": "ACCEPTED_EXCEPTION", "comment": "Approved for this cycle"},
    )
    assert decision.status_code == 200
    assert decision.json()["status"] == "ACCEPTED_EXCEPTION"
    country_anomalies = client.get(f"/api/countries/{country['id']}/anomalies")
    assert any(item["id"] == anomaly["id"] for item in country_anomalies.json())
    dashboard = client.get("/api/anomalies/dashboard")
    assert dashboard.status_code == 200
    assert dashboard.json()["totalAnomalies"] >= 1


def test_consolidation_download_special_country_and_name_collisions(client):
    source, template, _, _ = upload_template_and_mapping(client)
    countries = [
        create_country(client, template["id"], "Côte d'Ivoire / North", "CI"),
        create_country(client, template["id"], "Cote d Ivoire North", "CI"),
    ]
    for country in countries:
        imported = upload_country(client, country["id"], source)
        detail = analyze_version(client, imported["currentVersion"]["id"])
        assert detail["job"]["status"] == "COMPLETED"
    response = client.post(
        "/api/consolidations",
        json={"countryIds": [item["id"] for item in countries]},
    )
    assert response.status_code == 201, response.text
    job = response.json()
    assert job["status"] == "COMPLETED", job
    assert len(job["report"]["countriesIncluded"]) == 2
    download = client.get(f"/api/consolidations/{job['id']}/download")
    assert download.status_code == 200
    workbook = load_workbook(io.BytesIO(download.content), data_only=False)
    assert len(workbook.sheetnames) == 6
    assert len(set(name.casefold() for name in workbook.sheetnames)) == 6
    assert all(len(name) <= 31 for name in workbook.sheetnames)
    assert any(name.endswith("_2") for name in workbook.sheetnames)
    workbook.close()


def test_corrupted_and_unsupported_uploads_are_rejected(client):
    unsupported = client.post(
        "/api/templates",
        files={"file": ("template.xls", b"legacy", "application/vnd.ms-excel")},
    )
    assert unsupported.status_code == 415
    assert unsupported.json()["code"] == "UNSUPPORTED_FORMAT"
    corrupted = client.post(
        "/api/templates",
        files={"file": ("template.xlsx", b"PK\x03\x04broken", "application/octet-stream")},
    )
    assert corrupted.status_code == 422
    assert corrupted.json()["code"] == "FILE_CORRUPTED"
