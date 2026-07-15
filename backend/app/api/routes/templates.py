from __future__ import annotations

from typing import Annotated, Any

from fastapi import APIRouter, Depends, File, Form, Query, UploadFile, status
from openpyxl.utils import get_column_letter, range_boundaries
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, selectinload

from ...audit import record_audit
from ...config import Settings, get_settings
from ...database import get_db
from ...dependencies import Principal, get_principal
from ...errors import DomainError
from ...models import (
    SheetDefinition,
    TableColumnDefinition,
    TableDefinition,
    TableStatus,
    Template,
    TemplateVersion,
)
from ...schemas import (
    DetectTablesInput,
    MappingExport,
    MappingProgress,
    PreviewInput,
    SheetGrid,
    SheetIgnoreInput,
    SheetOut,
    StructuredPreview,
    TableCandidate,
    TableDefinitionCreate,
    TableDefinitionOut,
    TableDefinitionUpdate,
    TemplateOut,
    TemplateSummary,
)
from ...services.excel_reader import normalize_text
from ...services.runtime import (
    get_security_validator,
    get_storage,
    get_table_detector,
    get_workbook_reader,
    read_upload_limited,
)
from ..ownership import owned_template

router = APIRouter(prefix="/templates", tags=["templates"])


def _latest_version(template: Template) -> TemplateVersion:
    if not template.versions:
        raise DomainError("TEMPLATE_WITHOUT_VERSION", "Template has no imported version")
    return max(template.versions, key=lambda item: item.version)


async def _import_template(
    *,
    db: Session,
    principal: Principal,
    settings: Settings,
    upload: UploadFile,
    name: str | None,
    template: Template | None,
) -> Template:
    filename = upload.filename or "workbook.xlsx"
    data = await read_upload_limited(upload, settings)
    security = get_security_validator().validate(filename, data)
    storage = get_storage()
    key = storage.put_bytes(principal.organization.id, "templates", "xlsx", data)
    try:
        metadata = get_workbook_reader().inspect_workbook(storage.resolve(key))
    except Exception:
        storage.delete(key)
        raise
    metadata["security"] = {
        "zipEntries": security.zip_entries,
        "uncompressedBytes": security.uncompressed_bytes,
        "warnings": security.warnings,
    }
    if template is None:
        template_name = (name or filename.rsplit(".", 1)[0]).strip()
        if not template_name:
            raise DomainError("INVALID_TEMPLATE_NAME", "Template name is required")
        template = Template(
            organization_id=principal.organization.id,
            name=template_name[:255],
            latest_version=0,
            created_by_id=principal.user.id,
        )
        db.add(template)
        try:
            db.flush()
        except IntegrityError as exc:
            db.rollback()
            storage.delete(key)
            raise DomainError(
                "TEMPLATE_NAME_EXISTS",
                "A template with this name already exists in the organization",
                status_code=409,
            ) from exc
    elif any(version.sha256 == security.sha256 for version in template.versions):
        storage.delete(key)
        raise DomainError(
            "DUPLICATE_TEMPLATE_VERSION",
            "This exact workbook version has already been imported",
            status_code=409,
        )

    version_number = template.latest_version + 1
    version = TemplateVersion(
        template_id=template.id,
        version=version_number,
        mapping_version=1,
        original_filename=filename[:255],
        stored_key=key,
        sha256=security.sha256,
        size_bytes=security.size_bytes,
        sheet_count=metadata["sheetCount"],
        workbook_metadata=metadata,
        imported_by_id=principal.user.id,
    )
    db.add(version)
    db.flush()
    for sheet_metadata in metadata["sheets"]:
        db.add(
            SheetDefinition(
                template_version_id=version.id,
                name=sheet_metadata["name"],
                original_index=sheet_metadata["originalIndex"],
                visibility=sheet_metadata["visibility"],
                max_row=sheet_metadata["maxRow"],
                max_column=sheet_metadata["maxColumn"],
                merged_ranges=sheet_metadata["mergedRanges"],
                formula_cells=sheet_metadata["formulaCells"],
                native_tables=sheet_metadata["nativeTables"],
                named_ranges=sheet_metadata["namedRanges"],
                structural_signature=sheet_metadata["structuralSignature"],
            )
        )
    template.latest_version = version_number
    record_audit(
        db,
        principal,
        "TEMPLATE_IMPORTED",
        "TemplateVersion",
        version.id,
        {
            "templateId": template.id,
            "version": version_number,
            "sha256": security.sha256,
            "filename": filename,
        },
    )
    db.commit()
    return owned_template(db, principal, template.id)


@router.post("", response_model=TemplateOut, status_code=status.HTTP_201_CREATED)
async def create_template(
    file: Annotated[UploadFile, File(...)],
    name: Annotated[str | None, Form()] = None,
    db: Session = Depends(get_db),
    principal: Principal = Depends(get_principal),
    settings: Settings = Depends(get_settings),
) -> Template:
    return await _import_template(
        db=db,
        principal=principal,
        settings=settings,
        upload=file,
        name=name,
        template=None,
    )


@router.post(
    "/{template_id}/versions", response_model=TemplateOut, status_code=status.HTTP_201_CREATED
)
async def create_template_version(
    template_id: str,
    file: Annotated[UploadFile, File(...)],
    db: Session = Depends(get_db),
    principal: Principal = Depends(get_principal),
    settings: Settings = Depends(get_settings),
) -> Template:
    template = owned_template(db, principal, template_id)
    return await _import_template(
        db=db,
        principal=principal,
        settings=settings,
        upload=file,
        name=None,
        template=template,
    )


@router.get("", response_model=list[TemplateSummary])
def list_templates(
    db: Session = Depends(get_db), principal: Principal = Depends(get_principal)
) -> list[dict[str, Any]]:
    templates = list(
        db.scalars(
            select(Template)
            .where(Template.organization_id == principal.organization.id)
            .options(
                selectinload(Template.versions)
                .selectinload(TemplateVersion.sheets)
                .selectinload(SheetDefinition.tables)
            )
            .order_by(Template.updated_at.desc())
        )
    )
    result: list[dict[str, Any]] = []
    for template in templates:
        latest = _latest_version(template)
        result.append(
            {
                "id": template.id,
                "name": template.name,
                "latestVersion": template.latest_version,
                "originalFilename": latest.original_filename,
                "sha256": latest.sha256,
                "sheetCount": latest.sheet_count,
                "configuredSheets": sum(
                    sheet.mapping_status == "CONFIGURED" for sheet in latest.sheets
                ),
                "tableCount": sum(len(sheet.tables) for sheet in latest.sheets),
                "status": (
                    "READY"
                    if latest.sheets
                    and all(sheet.mapping_status == "CONFIGURED" for sheet in latest.sheets)
                    else (
                        "MAPPING"
                        if any(
                            sheet.mapping_status != "PENDING" or sheet.tables
                            for sheet in latest.sheets
                        )
                        else "DRAFT"
                    )
                ),
                "importedAt": latest.imported_at,
                "createdAt": template.created_at,
                "updatedAt": template.updated_at,
            }
        )
    return result


@router.get("/{template_id}", response_model=TemplateOut)
def get_template(
    template_id: str,
    db: Session = Depends(get_db),
    principal: Principal = Depends(get_principal),
) -> Template:
    return owned_template(db, principal, template_id)


@router.get("/{template_id}/sheets", response_model=list[SheetOut])
def list_sheets(
    template_id: str,
    version: int | None = Query(default=None, ge=1),
    db: Session = Depends(get_db),
    principal: Principal = Depends(get_principal),
) -> list[SheetDefinition]:
    template = owned_template(db, principal, template_id)
    selected = (
        next((item for item in template.versions if item.version == version), None)
        if version
        else _latest_version(template)
    )
    if selected is None:
        raise DomainError("TEMPLATE_VERSION_NOT_FOUND", "Template version not found", status_code=404)
    return selected.sheets


@router.get("/{template_id}/sheets/{sheet_id}/grid", response_model=SheetGrid)
def get_sheet_grid(
    template_id: str,
    sheet_id: str,
    range_ref: str | None = Query(default=None, alias="rangeRef"),
    legacy_range: str | None = Query(default=None, alias="range", include_in_schema=False),
    db: Session = Depends(get_db),
    principal: Principal = Depends(get_principal),
) -> dict[str, Any]:
    template = owned_template(db, principal, template_id)
    sheet = next(
        (
            sheet
            for version in template.versions
            for sheet in version.sheets
            if sheet.id == sheet_id
        ),
        None,
    )
    if sheet is None:
        raise DomainError("SHEET_NOT_FOUND", "Worksheet not found", status_code=404)
    version = sheet.template_version
    grid = get_workbook_reader().get_grid(
        get_storage().resolve(version.stored_key), sheet.name, range_ref or legacy_range
    )
    grid["sheetId"] = sheet.id
    return grid


@router.post("/{template_id}/tables/detect", response_model=list[TableCandidate])
def detect_tables(
    template_id: str,
    body: DetectTablesInput,
    db: Session = Depends(get_db),
    principal: Principal = Depends(get_principal),
) -> list[dict[str, Any]]:
    template = owned_template(db, principal, template_id)
    version = _latest_version(template)
    sheets = version.sheets
    if body.sheet_id:
        sheets = [sheet for sheet in sheets if sheet.id == body.sheet_id]
        if not sheets:
            raise DomainError("SHEET_NOT_FOUND", "Worksheet not found", status_code=404)
    candidates = get_table_detector().detect_candidates(
        get_storage().resolve(version.stored_key), [sheet.name for sheet in sheets]
    )
    sheet_ids = {sheet.name: sheet.id for sheet in sheets}
    result = []
    for candidate in candidates:
        if candidate["confidence"] < body.minimum_confidence:
            continue
        candidate["sheetId"] = sheet_ids[candidate["sheetName"]]
        result.append(candidate)
    return result


@router.post("/{template_id}/tables/preview", response_model=StructuredPreview)
def preview_table(
    template_id: str,
    body: PreviewInput,
    db: Session = Depends(get_db),
    principal: Principal = Depends(get_principal),
) -> dict[str, Any]:
    template = owned_template(db, principal, template_id)
    version = _latest_version(template)
    sheet = next((item for item in version.sheets if item.id == body.sheet_id), None)
    if sheet is None:
        raise DomainError("SHEET_NOT_FOUND", "Worksheet not found", status_code=404)
    return get_workbook_reader().preview_range(
        get_storage().resolve(version.stored_key),
        sheet.name,
        body.range_ref,
        body.header_rows,
        body.data_start_row,
        body.data_end_row,
        body.key_columns,
    )


def _owned_sheet(
    db: Session, principal: Principal, template_id: str, sheet_id: str
) -> tuple[Template, TemplateVersion, SheetDefinition]:
    template = owned_template(db, principal, template_id)
    for version in template.versions:
        for sheet in version.sheets:
            if sheet.id == sheet_id:
                return template, version, sheet
    raise DomainError("SHEET_NOT_FOUND", "Worksheet not found", status_code=404)


def _validate_definition_bounds(body: Any) -> tuple[int, int, int, int]:
    min_col, min_row, max_col, max_row = range_boundaries(body.range_ref)
    if any(row < min_row or row > max_row for row in body.header_rows):
        raise DomainError("INVALID_HEADER_ROWS", "Header rows must be inside the selected range")
    if not min_row <= body.data_start_row <= max_row:
        raise DomainError("INVALID_DATA_START", "Data start row must be inside the selected range")
    if body.data_end_row is not None and not min_row <= body.data_end_row <= max_row:
        raise DomainError("INVALID_DATA_END", "Data end row must be inside the selected range")
    allowed_columns = {get_column_letter(column) for column in range(min_col, max_col + 1)}
    configured_columns = set(
        body.key_columns + body.value_columns + body.computed_columns + body.ignored_columns
    )
    if not configured_columns <= allowed_columns:
        raise DomainError(
            "COLUMN_OUTSIDE_RANGE", "All configured columns must be inside the selected range"
        )
    return min_col, min_row, max_col, max_row


def _replace_columns(
    db: Session,
    table: TableDefinition,
    inputs: list[Any],
    signature: dict[str, Any],
    min_col: int,
    max_col: int,
) -> None:
    table.columns.clear()
    allowed = {get_column_letter(column) for column in range(min_col, max_col + 1)}
    if inputs:
        ordinals = [item.ordinal for item in inputs]
        if len(ordinals) != len(set(ordinals)):
            raise DomainError("DUPLICATE_COLUMN_ORDINAL", "Column ordinals must be unique")
        for item in inputs:
            if item.excel_column not in allowed:
                raise DomainError("COLUMN_OUTSIDE_RANGE", "A column definition is outside the range")
            table.columns.append(
                TableColumnDefinition(
                    excel_column=item.excel_column,
                    name=item.name,
                    normalized_name=normalize_text(item.name),
                    data_type=item.data_type,
                    role=item.role,
                    ordinal=item.ordinal,
                    required=item.required,
                )
            )
        return
    headers = signature.get("headers", [])
    for ordinal, column in enumerate(range(min_col, max_col + 1)):
        letter = get_column_letter(column)
        role = "VALUE"
        if letter in table.key_columns:
            role = "KEY"
        elif letter in table.computed_columns:
            role = "CALCULATED"
        elif letter in table.ignored_columns:
            role = "IGNORE"
        name = headers[ordinal] if ordinal < len(headers) else letter
        table.columns.append(
            TableColumnDefinition(
                excel_column=letter,
                name=name,
                normalized_name=normalize_text(name),
                role=role,
                ordinal=ordinal,
                required=letter not in table.variable_columns,
            )
        )


@router.post(
    "/{template_id}/tables",
    response_model=TableDefinitionOut,
    status_code=status.HTTP_201_CREATED,
)
def create_table_definition(
    template_id: str,
    body: TableDefinitionCreate,
    db: Session = Depends(get_db),
    principal: Principal = Depends(get_principal),
) -> TableDefinition:
    template, version, sheet = _owned_sheet(db, principal, template_id, body.sheet_id)
    if version.version != template.latest_version:
        raise DomainError(
            "HISTORICAL_MAPPING_READ_ONLY",
            "Only the latest template version can be mapped",
            status_code=409,
        )
    min_col, _, max_col, _ = _validate_definition_bounds(body)
    table = TableDefinition(
        sheet_definition_id=sheet.id,
        name=body.name,
        range_ref=body.range_ref,
        header_rows=body.header_rows,
        data_start_row=body.data_start_row,
        data_end_row=body.data_end_row,
        data_end_rule=body.data_end_rule,
        key_columns=body.key_columns,
        value_columns=body.value_columns,
        total_rows=body.total_rows,
        computed_columns=body.computed_columns,
        structure_mode=(
            body.structure_mode.value
            if hasattr(body.structure_mode, "value")
            else body.structure_mode
        ),
        required=body.required,
        variable_rows=body.variable_rows,
        variable_columns=body.variable_columns,
        ignored_rows=body.ignored_rows,
        ignored_columns=body.ignored_columns,
        required_cells=body.required_cells,
        required_formulas=body.required_formulas,
        tolerate_blank_rows_columns=body.tolerate_blank_rows_columns,
        orientation=body.orientation,
        status=TableStatus.DRAFT.value,
        created_by_id=principal.user.id,
    )
    db.add(table)
    db.flush()
    table.signature = get_workbook_reader().build_table_signature(
        get_storage().resolve(version.stored_key), sheet.name, table
    )
    _replace_columns(db, table, body.columns, table.signature, min_col, max_col)
    sheet.mapping_status = "IN_PROGRESS"
    version.mapping_version += 1
    record_audit(
        db,
        principal,
        "TABLE_DEFINITION_CREATED",
        "TableDefinition",
        table.id,
        {"templateId": template.id, "sheet": sheet.name, "rangeRef": table.range_ref},
    )
    db.commit()
    return table


def _owned_table(
    db: Session, principal: Principal, template_id: str, table_id: str
) -> tuple[Template, TemplateVersion, SheetDefinition, TableDefinition]:
    template = owned_template(db, principal, template_id)
    for version in template.versions:
        for sheet in version.sheets:
            for table in sheet.tables:
                if table.id == table_id:
                    return template, version, sheet, table
    raise DomainError("TABLE_NOT_FOUND", "Table definition not found", status_code=404)


@router.put("/{template_id}/tables/{table_id}", response_model=TableDefinitionOut)
def update_table_definition(
    template_id: str,
    table_id: str,
    body: TableDefinitionUpdate,
    db: Session = Depends(get_db),
    principal: Principal = Depends(get_principal),
) -> TableDefinition:
    template, version, sheet, table = _owned_table(db, principal, template_id, table_id)
    if version.version != template.latest_version:
        raise DomainError("HISTORICAL_MAPPING_READ_ONLY", "Historical mappings are read-only", status_code=409)
    changes = body.model_dump(exclude_unset=True)
    column_inputs = changes.pop("columns", None)
    for field, value in changes.items():
        if field == "structure_mode" and value is not None:
            value = value.value if hasattr(value, "value") else value
        setattr(table, field, value)
    # Revalidate the effective definition after partial updates.
    effective = TableDefinitionCreate(
        sheet_id=sheet.id,
        name=table.name,
        range_ref=table.range_ref,
        header_rows=table.header_rows,
        data_start_row=table.data_start_row,
        data_end_row=table.data_end_row,
        data_end_rule=table.data_end_rule,
        key_columns=table.key_columns,
        value_columns=table.value_columns,
        total_rows=table.total_rows,
        computed_columns=table.computed_columns,
        structure_mode=table.structure_mode,
        required=table.required,
        variable_rows=table.variable_rows,
        variable_columns=table.variable_columns,
        ignored_rows=table.ignored_rows,
        ignored_columns=table.ignored_columns,
        required_cells=table.required_cells,
        required_formulas=table.required_formulas,
        tolerate_blank_rows_columns=table.tolerate_blank_rows_columns,
        orientation=table.orientation,
    )
    min_col, _, max_col, _ = _validate_definition_bounds(effective)
    table.signature = get_workbook_reader().build_table_signature(
        get_storage().resolve(version.stored_key), sheet.name, table
    )
    if column_inputs is not None:
        _replace_columns(db, table, column_inputs, table.signature, min_col, max_col)
    elif body.range_ref is not None:
        _replace_columns(db, table, [], table.signature, min_col, max_col)
    table.status = TableStatus.DRAFT.value
    sheet.mapping_status = "IN_PROGRESS"
    version.mapping_version += 1
    record_audit(
        db,
        principal,
        "TABLE_DEFINITION_UPDATED",
        "TableDefinition",
        table.id,
        {"changedFields": sorted(changes)},
    )
    db.commit()
    return table


@router.post("/{template_id}/tables/{table_id}/validate", response_model=TableDefinitionOut)
def validate_table_definition(
    template_id: str,
    table_id: str,
    db: Session = Depends(get_db),
    principal: Principal = Depends(get_principal),
) -> TableDefinition:
    _, version, sheet, table = _owned_table(db, principal, template_id, table_id)
    if not table.signature.get("headers"):
        raise DomainError("TABLE_SIGNATURE_EMPTY", "The table signature is incomplete")
    table.status = TableStatus.VALIDATED.value
    if all(item.status == TableStatus.VALIDATED.value for item in sheet.tables):
        sheet.mapping_status = "CONFIGURED"
    version.mapping_version += 1
    record_audit(
        db, principal, "TABLE_DEFINITION_VALIDATED", "TableDefinition", table.id
    )
    db.commit()
    return table


@router.post("/{template_id}/tables/{table_id}/reject", response_model=TableDefinitionOut)
def reject_table_definition(
    template_id: str,
    table_id: str,
    db: Session = Depends(get_db),
    principal: Principal = Depends(get_principal),
) -> TableDefinition:
    _, version, sheet, table = _owned_table(db, principal, template_id, table_id)
    table.status = TableStatus.REJECTED.value
    sheet.mapping_status = "IN_PROGRESS"
    version.mapping_version += 1
    record_audit(db, principal, "TABLE_DEFINITION_REJECTED", "TableDefinition", table.id)
    db.commit()
    return table


@router.patch("/{template_id}/sheets/{sheet_id}", response_model=SheetOut)
def set_sheet_ignored(
    template_id: str,
    sheet_id: str,
    body: SheetIgnoreInput,
    db: Session = Depends(get_db),
    principal: Principal = Depends(get_principal),
) -> SheetDefinition:
    _, version, sheet = _owned_sheet(db, principal, template_id, sheet_id)
    sheet.ignored = body.ignored
    sheet.mapping_status = "CONFIGURED" if body.ignored else "PENDING"
    version.mapping_version += 1
    record_audit(
        db,
        principal,
        "SHEET_MAPPING_STATUS_CHANGED",
        "SheetDefinition",
        sheet.id,
        {"ignored": body.ignored},
    )
    db.commit()
    return sheet


@router.get("/{template_id}/mapping/progress", response_model=MappingProgress)
def mapping_progress(
    template_id: str,
    db: Session = Depends(get_db),
    principal: Principal = Depends(get_principal),
) -> dict[str, int]:
    template = owned_template(db, principal, template_id)
    version = _latest_version(template)
    total = len(version.sheets)
    configured = sum(sheet.mapping_status == "CONFIGURED" for sheet in version.sheets)
    validated = sum(
        table.status == TableStatus.VALIDATED.value
        for sheet in version.sheets
        for table in sheet.tables
    )
    return {
        "detectedSheets": total,
        "configuredSheets": configured,
        "remainingSheets": total - configured,
        "validatedTables": validated,
        "percent": round(configured / max(1, total) * 100),
    }


@router.get("/{template_id}/mapping", response_model=MappingExport)
def export_mapping(
    template_id: str,
    db: Session = Depends(get_db),
    principal: Principal = Depends(get_principal),
) -> dict[str, Any]:
    template = owned_template(db, principal, template_id)
    version = _latest_version(template)
    return {
        "templateId": template.id,
        "templateVersion": version.version,
        "mappingVersion": version.mapping_version,
        "workbookHash": version.sha256,
        "sheets": [
            {
                "id": sheet.id,
                "name": sheet.name,
                "originalIndex": sheet.original_index,
                "ignored": sheet.ignored,
                "status": sheet.mapping_status,
                "tables": [TableDefinitionOut.model_validate(table).model_dump(by_alias=True) for table in sheet.tables],
            }
            for sheet in version.sheets
        ],
    }
