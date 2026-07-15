from __future__ import annotations

import hashlib
import json
import math
import re
from collections import Counter
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import (
    column_index_from_string,
    get_column_letter,
    range_boundaries,
)
from openpyxl.worksheet.worksheet import Worksheet

from ..errors import DomainError


def json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, bool)):
        return value
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return str(value)
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, timedelta):
        return str(value)
    return str(value)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = re.sub(r"\s+", " ", str(value)).strip().casefold()
    return re.sub(r"[^\w]+", " ", text, flags=re.UNICODE).strip()


def estimate_type(values: Iterable[Any]) -> str:
    observed: Counter[str] = Counter()
    for value in values:
        if value is None or value == "":
            continue
        if isinstance(value, str) and value.startswith("="):
            observed["formula"] += 1
        elif isinstance(value, bool):
            observed["boolean"] += 1
        elif isinstance(value, (int, float, Decimal)):
            observed["number"] += 1
        elif isinstance(value, (datetime, date, time)):
            observed["date"] += 1
        else:
            observed["text"] += 1
    if not observed:
        return "empty"
    if len(observed) == 1:
        return next(iter(observed))
    return "mixed"


def _color_value(color: Any) -> str | None:
    if color is None:
        return None
    for attribute in ("rgb", "indexed", "theme", "auto"):
        try:
            value = getattr(color, attribute, None)
        except (TypeError, ValueError):
            continue
        if value is not None and not hasattr(value, "type"):
            return str(value)
    return None


def cell_style(cell: Any) -> dict[str, Any]:
    if isinstance(cell, MergedCell):
        return {}
    border = cell.border
    return {
        "styleId": cell.style_id,
        "bold": bool(cell.font and cell.font.bold),
        "italic": bool(cell.font and cell.font.italic),
        "fill": _color_value(cell.fill.fgColor) if cell.fill else None,
        "fontColor": _color_value(cell.font.color) if cell.font else None,
        "horizontalAlignment": cell.alignment.horizontal if cell.alignment else None,
        "verticalAlignment": cell.alignment.vertical if cell.alignment else None,
        "hasBorder": any(
            getattr(border, side).style is not None
            for side in ("left", "right", "top", "bottom")
        ),
    }


def safe_load_workbook(path: Path):
    try:
        # data_only=False guarantees formulas remain expressions. openpyxl never evaluates them.
        return load_workbook(
            filename=path,
            read_only=False,
            data_only=False,
            keep_vba=False,
            keep_links=False,
        )
    except Exception as exc:  # openpyxl raises several parser-specific exception types
        raise DomainError(
            "FILE_CORRUPTED",
            "The workbook cannot be read safely",
            status_code=422,
        ) from exc


class OpenpyxlWorkbookReader:
    def __init__(self, max_grid_cells: int = 5_000) -> None:
        self.max_grid_cells = max_grid_cells

    def inspect_workbook(self, file_path: Path) -> dict[str, Any]:
        workbook = safe_load_workbook(file_path)
        try:
            if not workbook.worksheets:
                raise DomainError("WORKBOOK_WITHOUT_SHEETS", "The workbook has no worksheets")

            defined_names = self._defined_names(workbook)
            sheets: list[dict[str, Any]] = []
            total_formula_count = 0
            for index, worksheet in enumerate(workbook.worksheets):
                formula_cells: list[dict[str, Any]] = []
                non_empty = 0
                for row in worksheet.iter_rows():
                    for cell in row:
                        if isinstance(cell, MergedCell):
                            continue
                        if cell.value is not None:
                            non_empty += 1
                        if cell.data_type == "f" or (
                            isinstance(cell.value, str) and cell.value.startswith("=")
                        ):
                            formula_cells.append(
                                {"coordinate": cell.coordinate, "formula": str(cell.value)}
                            )
                total_formula_count += len(formula_cells)

                native_tables: list[dict[str, Any]] = []
                for name in worksheet.tables:
                    table = worksheet.tables[name]
                    native_tables.append(
                        {
                            "name": table.name,
                            "displayName": table.displayName,
                            "rangeRef": table.ref.replace("$", ""),
                        }
                    )

                sheet_names = [
                    item
                    for item in defined_names
                    if any(destination[0] == worksheet.title for destination in item["destinations"])
                ]
                merged = [str(value) for value in worksheet.merged_cells.ranges]
                signature_payload = {
                    "name": worksheet.title,
                    "index": index,
                    "visibility": worksheet.sheet_state,
                    "dimensions": worksheet.calculate_dimension(),
                    "merged": merged,
                    "nativeTables": native_tables,
                    "labels": self._sample_labels(worksheet),
                }
                sheets.append(
                    {
                        "name": worksheet.title,
                        "originalIndex": index,
                        "visibility": worksheet.sheet_state,
                        "maxRow": worksheet.max_row,
                        "maxColumn": worksheet.max_column,
                        "dimension": worksheet.calculate_dimension(),
                        "nonEmptyCells": non_empty,
                        "mergedRanges": merged,
                        "formulaCount": len(formula_cells),
                        "formulaCells": formula_cells,
                        "nativeTables": native_tables,
                        "namedRanges": sheet_names,
                        "structuralSignature": {
                            "sha256": hashlib.sha256(
                                json.dumps(
                                    signature_payload,
                                    sort_keys=True,
                                    ensure_ascii=False,
                                ).encode("utf-8")
                            ).hexdigest(),
                            "labels": signature_payload["labels"],
                        },
                    }
                )
            return {
                "sheetCount": len(sheets),
                "sheetNames": [sheet["name"] for sheet in sheets],
                "sheets": sheets,
                "definedNames": defined_names,
                "formulaCount": total_formula_count,
                "calculationMode": getattr(workbook.calculation, "calcMode", None),
                "formulasEvaluated": False,
            }
        finally:
            workbook.close()

    def get_grid(
        self,
        file_path: Path,
        sheet_name: str,
        range_ref: str | None = None,
    ) -> dict[str, Any]:
        workbook = safe_load_workbook(file_path)
        try:
            if sheet_name not in workbook.sheetnames:
                raise DomainError("SHEET_NOT_FOUND", "Worksheet not found", status_code=404)
            worksheet = workbook[sheet_name]
            if range_ref:
                min_col, min_row, max_col, max_row = self._range_bounds(range_ref)
            else:
                min_col, min_row = 1, 1
                max_col, max_row = worksheet.max_column, worksheet.max_row
            cells_requested = (max_row - min_row + 1) * (max_col - min_col + 1)
            if cells_requested > self.max_grid_cells:
                raise DomainError(
                    "GRID_RANGE_TOO_LARGE",
                    "Requested grid range exceeds the configured cell limit",
                    status_code=422,
                    details={"maximumCells": self.max_grid_cells},
                )
            merged_lookup = self._merged_lookup(worksheet, min_row, max_row, min_col, max_col)
            cells: list[dict[str, Any]] = []
            for row in worksheet.iter_rows(
                min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
            ):
                for cell in row:
                    raw = cell.value
                    formula = str(raw) if isinstance(raw, str) and raw.startswith("=") else None
                    cells.append(
                        {
                            "coordinate": cell.coordinate,
                            "row": cell.row,
                            "column": cell.column,
                            "value": json_value(raw),
                            "formula": formula,
                            "dataType": getattr(cell, "data_type", None),
                            "numberFormat": getattr(cell, "number_format", None),
                            "style": cell_style(cell),
                            "mergedRange": merged_lookup.get(cell.coordinate),
                        }
                    )
            return {
                "sheetName": worksheet.title,
                "rangeRef": (
                    f"{get_column_letter(min_col)}{min_row}:"
                    f"{get_column_letter(max_col)}{max_row}"
                ),
                "minRow": min_row,
                "maxRow": max_row,
                "minColumn": min_col,
                "maxColumn": max_col,
                "cells": cells,
                "mergedRanges": sorted(set(merged_lookup.values())),
            }
        finally:
            workbook.close()

    def preview_range(
        self,
        file_path: Path,
        sheet_name: str,
        range_ref: str,
        header_rows: list[int],
        data_start_row: int,
        data_end_row: int | None,
        key_columns: list[str],
    ) -> dict[str, Any]:
        workbook = safe_load_workbook(file_path)
        try:
            if sheet_name not in workbook.sheetnames:
                raise DomainError("SHEET_NOT_FOUND", "Worksheet not found", status_code=404)
            ws = workbook[sheet_name]
            min_col, min_row, max_col, range_max_row = self._range_bounds(range_ref)
            max_row = min(data_end_row or range_max_row, range_max_row)
            headers = self._combined_headers(ws, header_rows, min_col, max_col)
            unique_headers = self._unique_headers(headers)
            column_values: list[list[Any]] = [[] for _ in unique_headers]
            rows: list[dict[str, Any]] = []
            coordinates: list[list[str]] = []
            formula_cells: list[dict[str, Any]] = []
            empty_cells: list[str] = []
            total_rows: list[int] = []
            key_indexes = [
                column_index_from_string(column) - min_col
                for column in key_columns
                if min_col <= column_index_from_string(column) <= max_col
            ]
            seen_keys: dict[tuple[str, ...], int] = {}
            duplicates: list[dict[str, Any]] = []
            for row_number in range(data_start_row, max_row + 1):
                values: list[Any] = []
                row_coordinates: list[str] = []
                for offset, column_number in enumerate(range(min_col, max_col + 1)):
                    cell = ws.cell(row=row_number, column=column_number)
                    value = json_value(cell.value)
                    values.append(value)
                    column_values[offset].append(cell.value)
                    row_coordinates.append(cell.coordinate)
                    if value in (None, "") and len(empty_cells) < 500:
                        empty_cells.append(cell.coordinate)
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_cells.append(
                            {"coordinate": cell.coordinate, "formula": cell.value}
                        )
                coordinates.append(row_coordinates)
                rows.append(dict(zip(unique_headers, values, strict=True)))
                first_labels = " ".join(normalize_text(value) for value in values[:2])
                if re.search(r"\b(total|totaux|subtotal|sous total|overall|grand total)\b", first_labels):
                    total_rows.append(row_number)
                if key_indexes:
                    key = tuple(normalize_text(values[index]) for index in key_indexes)
                    if any(key):
                        if key in seen_keys:
                            duplicates.append(
                                {"key": list(key), "rows": [seen_keys[key], row_number]}
                            )
                        else:
                            seen_keys[key] = row_number
            return {
                "sheetName": sheet_name,
                "rangeRef": range_ref,
                "headers": unique_headers,
                "estimatedTypes": [estimate_type(values) for values in column_values],
                "rows": rows[:100],
                "coordinates": coordinates[:100],
                "formulaCells": formula_cells,
                "emptyCells": empty_cells,
                "duplicateKeys": duplicates,
                "detectedTotalRows": total_rows,
            }
        finally:
            workbook.close()

    def build_table_signature(
        self,
        file_path: Path,
        sheet_name: str,
        definition: Any,
    ) -> dict[str, Any]:
        workbook = safe_load_workbook(file_path)
        try:
            ws = workbook[sheet_name]
            range_ref = definition.range_ref
            min_col, min_row, max_col, max_row = self._range_bounds(range_ref)
            header_rows = list(definition.header_rows)
            headers = self._combined_headers(ws, header_rows, min_col, max_col)
            headers = self._unique_headers(headers)
            data_end_row = min(definition.data_end_row or max_row, max_row)
            key_columns = list(definition.key_columns)
            key_col_indexes = [
                column_index_from_string(column)
                for column in key_columns
                if min_col <= column_index_from_string(column) <= max_col
            ]
            row_keys: list[dict[str, Any]] = []
            total_rows = set(definition.total_rows)
            for row_number in range(definition.data_start_row, data_end_row + 1):
                if row_number in total_rows:
                    continue
                values = [json_value(ws.cell(row_number, col).value) for col in key_col_indexes]
                if any(value not in (None, "") for value in values):
                    row_keys.append({"row": row_number, "values": values})

            formulas: list[dict[str, Any]] = []
            for row in ws.iter_rows(
                min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
            ):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formulas.append(
                            {
                                "coordinate": cell.coordinate,
                                "relativeRow": cell.row - min_row,
                                "relativeColumn": cell.column - min_col,
                                "formula": cell.value,
                            }
                        )
            required_cells = []
            for item in definition.required_cells:
                coordinate = str(item.get("coordinate", "")).upper().replace("$", "")
                if coordinate:
                    required_cells.append(
                        {
                            **item,
                            "coordinate": coordinate,
                            "expected": json_value(ws[coordinate].value),
                        }
                    )
            required_formulas = []
            for item in definition.required_formulas:
                coordinate = str(item.get("coordinate", "")).upper().replace("$", "")
                if coordinate:
                    required_formulas.append(
                        {
                            **item,
                            "coordinate": coordinate,
                            "formula": str(ws[coordinate].value or ""),
                            "relativeRow": ws[coordinate].row - min_row,
                            "relativeColumn": ws[coordinate].column - min_col,
                        }
                    )
            preview_matrix: list[list[Any]] = []
            for row_number in range(min_row, min(max_row, min_row + 14) + 1):
                preview_matrix.append(
                    [json_value(ws.cell(row_number, col).value) for col in range(min_col, max_col + 1)]
                )
            payload = {
                "rangeRef": range_ref,
                "anchor": f"{get_column_letter(min_col)}{min_row}",
                "minRow": min_row,
                "minColumn": min_col,
                "width": max_col - min_col + 1,
                "height": max_row - min_row + 1,
                "headers": headers,
                "normalizedHeaders": [normalize_text(value) for value in headers],
                "headerRows": header_rows,
                "dataStartRow": definition.data_start_row,
                "dataEndRow": definition.data_end_row,
                "keyColumns": key_columns,
                "rowKeys": row_keys,
                "formulas": formulas,
                "requiredCells": required_cells,
                "requiredFormulas": required_formulas,
                "mergedRanges": [
                    str(merged)
                    for merged in ws.merged_cells.ranges
                    if not (
                        merged.max_col < min_col
                        or merged.min_col > max_col
                        or merged.max_row < min_row
                        or merged.min_row > max_row
                    )
                ],
                "preview": {"headers": headers, "rows": preview_matrix},
            }
            payload["sha256"] = hashlib.sha256(
                json.dumps(payload, sort_keys=True, ensure_ascii=False).encode("utf-8")
            ).hexdigest()
            return payload
        finally:
            workbook.close()

    @staticmethod
    def _range_bounds(range_ref: str) -> tuple[int, int, int, int]:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(
                range_ref.replace("$", "").upper()
            )
        except (TypeError, ValueError) as exc:
            raise DomainError("INVALID_RANGE", "Invalid Excel range") from exc
        if None in (min_col, min_row, max_col, max_row):
            raise DomainError("INVALID_RANGE", "A rectangular Excel range is required")
        return min_col, min_row, max_col, max_row

    @staticmethod
    def _combined_headers(
        ws: Worksheet, header_rows: list[int], min_col: int, max_col: int
    ) -> list[str]:
        headers: list[str] = []
        for column in range(min_col, max_col + 1):
            parts = [
                str(ws.cell(row=row_number, column=column).value).strip()
                for row_number in header_rows
                if ws.cell(row=row_number, column=column).value not in (None, "")
            ]
            headers.append(" / ".join(parts) or get_column_letter(column))
        return headers

    @staticmethod
    def _unique_headers(headers: list[str]) -> list[str]:
        counts: Counter[str] = Counter()
        result: list[str] = []
        for header in headers:
            counts[header] += 1
            result.append(header if counts[header] == 1 else f"{header}_{counts[header]}")
        return result

    @staticmethod
    def _defined_names(workbook: Any) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = []
        for definition in workbook.defined_names.values():
            destinations: list[list[str]] = []
            try:
                destinations = [[sheet, ref] for sheet, ref in definition.destinations]
            except (AttributeError, TypeError, ValueError):
                pass
            result.append(
                {
                    "name": definition.name,
                    "text": definition.attr_text,
                    "localSheetId": definition.localSheetId,
                    "hidden": bool(definition.hidden),
                    "destinations": destinations,
                }
            )
        return result

    @staticmethod
    def _sample_labels(ws: Worksheet, limit: int = 150) -> list[str]:
        labels: list[str] = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and not cell.value.startswith("="):
                    normalized = normalize_text(cell.value)
                    if normalized and normalized not in labels:
                        labels.append(normalized)
                        if len(labels) >= limit:
                            return labels
        return labels

    @staticmethod
    def _merged_lookup(
        ws: Worksheet, min_row: int, max_row: int, min_col: int, max_col: int
    ) -> dict[str, str]:
        lookup: dict[str, str] = {}
        for merged in ws.merged_cells.ranges:
            if (
                merged.max_row < min_row
                or merged.min_row > max_row
                or merged.max_col < min_col
                or merged.min_col > max_col
            ):
                continue
            for row in range(max(min_row, merged.min_row), min(max_row, merged.max_row) + 1):
                for col in range(max(min_col, merged.min_col), min(max_col, merged.max_col) + 1):
                    lookup[f"{get_column_letter(col)}{row}"] = str(merged)
        return lookup
