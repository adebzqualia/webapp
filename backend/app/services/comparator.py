from __future__ import annotations

from collections import Counter
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries

from ..models import SheetDefinition, StructureMode, TableDefinition, TemplateVersion
from .excel_reader import json_value, normalize_text, safe_load_workbook
from .table_detector import HeuristicTableDetector


def similarity(left: str, right: str) -> float:
    return SequenceMatcher(None, left.casefold(), right.casefold()).ratio()


class WorkbookStructureComparator:
    """Compares a country workbook without evaluating any formula."""

    def __init__(self, detector: HeuristicTableDetector | None = None) -> None:
        self.detector = detector or HeuristicTableDetector()

    def compare(
        self,
        template_version: TemplateVersion,
        country_path: Path,
        country_metadata: dict[str, Any],
    ) -> dict[str, Any]:
        workbook = safe_load_workbook(country_path)
        try:
            metadata_by_name = {
                sheet["name"]: sheet for sheet in country_metadata.get("sheets", [])
            }
            expected_sheets = sorted(template_version.sheets, key=lambda item: item.original_index)
            expected_names = [sheet.name for sheet in expected_sheets]
            actual_names = list(workbook.sheetnames)
            anomalies: list[dict[str, Any]] = []
            extracted: list[dict[str, Any]] = []

            matched: dict[str, str] = {
                name: name for name in expected_names if name in actual_names
            }
            missing = [sheet for sheet in expected_sheets if sheet.name not in matched]
            added = [name for name in actual_names if name not in matched.values()]
            rename_matches = self._match_renamed_sheets(missing, added, metadata_by_name)
            for sheet, actual_name, confidence, reasons in rename_matches:
                matched[sheet.name] = actual_name
                missing.remove(sheet)
                added.remove(actual_name)
                anomalies.append(
                    self._anomaly(
                        "SHEET_RENAMED",
                        "ERROR",
                        f"La feuille « {sheet.name} » semble avoir été renommée « {actual_name} ».",
                        sheet_name=actual_name,
                        expected=sheet.name,
                        actual=actual_name,
                        suggestion="Confirmer l’hypothèse de renommage ou restaurer le nom attendu.",
                        confidence=confidence,
                        match_reasons=reasons,
                    )
                )
            for sheet in missing:
                anomalies.append(
                    self._anomaly(
                        "SHEET_MISSING",
                        "BLOCKING",
                        f"La feuille obligatoire « {sheet.name} » est absente.",
                        sheet_name=sheet.name,
                        expected=sheet.name,
                        actual=None,
                        suggestion="Restaurer la feuille depuis le template de référence.",
                    )
                )
            for actual_name in added:
                anomalies.append(
                    self._anomaly(
                        "SHEET_ADDED",
                        "WARNING",
                        f"Une feuille supplémentaire « {actual_name} » a été détectée.",
                        sheet_name=actual_name,
                        expected=None,
                        actual=actual_name,
                        suggestion="Vérifier si cette feuille est intentionnelle.",
                    )
                )

            actual_expected_order = [
                expected
                for actual in actual_names
                for expected, matched_actual in matched.items()
                if actual == matched_actual
            ]
            comparable_expected = [name for name in expected_names if name in matched]
            if actual_expected_order != comparable_expected:
                anomalies.append(
                    self._anomaly(
                        "SHEET_ORDER_CHANGED",
                        "WARNING",
                        "L’ordre des feuilles diffère du template.",
                        expected=comparable_expected,
                        actual=actual_expected_order,
                        suggestion="Rétablir l’ordre du template si l’ordre est structurel.",
                    )
                )

            for sheet in expected_sheets:
                actual_name = matched.get(sheet.name)
                if actual_name is None:
                    continue
                actual_ws = workbook[actual_name]
                actual_meta = metadata_by_name.get(actual_name, {})
                if sheet.visibility != actual_ws.sheet_state:
                    anomalies.append(
                        self._anomaly(
                            "SHEET_VISIBILITY_CHANGED",
                            "WARNING",
                            f"La visibilité de la feuille « {actual_name} » a changé.",
                            sheet_name=actual_name,
                            expected=sheet.visibility,
                            actual=actual_ws.sheet_state,
                            suggestion="Vérifier que ce changement de visibilité est volontaire.",
                        )
                    )
                if self._major_dimension_change(sheet, actual_ws):
                    anomalies.append(
                        self._anomaly(
                            "SHEET_DIMENSIONS_CHANGED",
                            "WARNING",
                            f"Les dimensions utilisées de « {actual_name} » diffèrent fortement.",
                            sheet_name=actual_name,
                            expected={"rows": sheet.max_row, "columns": sheet.max_column},
                            actual={"rows": actual_ws.max_row, "columns": actual_ws.max_column},
                            suggestion="Contrôler les insertions ou suppressions de lignes et colonnes.",
                        )
                    )
                expected_merges = set(sheet.merged_ranges or [])
                actual_merges = {str(value) for value in actual_ws.merged_cells.ranges}
                if expected_merges != actual_merges:
                    anomalies.append(
                        self._anomaly(
                            "MERGED_CELL_CHANGED",
                            "WARNING",
                            f"Les cellules fusionnées de « {actual_name} » ont changé.",
                            sheet_name=actual_name,
                            expected=sorted(expected_merges),
                            actual=sorted(actual_merges),
                            suggestion="Comparer les zones fusionnées avec le template.",
                        )
                    )
                expected_ranges = self._named_range_map(sheet.named_ranges or [])
                actual_ranges = self._named_range_map(actual_meta.get("namedRanges", []))
                if expected_ranges != actual_ranges:
                    anomalies.append(
                        self._anomaly(
                            "NAMED_RANGE_CHANGED",
                            "WARNING",
                            f"Les plages nommées de « {actual_name} » ont changé.",
                            sheet_name=actual_name,
                            expected=expected_ranges,
                            actual=actual_ranges,
                            suggestion="Vérifier les plages nommées utilisées par le classeur.",
                        )
                    )

                if sheet.ignored:
                    continue
                candidates = self.detector.detect_candidates(country_path, [actual_name])
                for definition in sheet.tables:
                    if definition.status != "VALIDATED":
                        continue
                    table_anomalies, table_extract = self._compare_table(
                        actual_ws,
                        definition,
                        candidates,
                    )
                    anomalies.extend(table_anomalies)
                    if table_extract is not None:
                        extracted.append(table_extract)

            severity_counts = Counter(item["severity"] for item in anomalies)
            category_counts = Counter(item["category"] for item in anomalies)
            return {
                "anomalies": anomalies,
                "extractedTables": extracted,
                "summary": {
                    "anomalyCount": len(anomalies),
                    "blockingCount": severity_counts["BLOCKING"],
                    "errorCount": severity_counts["ERROR"],
                    "warningCount": severity_counts["WARNING"],
                    "informationCount": severity_counts["INFO"],
                    "categories": dict(category_counts),
                    "sheetsCompared": len(matched),
                    "tablesExtracted": len(extracted),
                    "formulasEvaluated": False,
                },
            }
        finally:
            workbook.close()

    def _compare_table(
        self,
        ws: Any,
        definition: TableDefinition,
        detected_candidates: list[dict[str, Any]],
    ) -> tuple[list[dict[str, Any]], dict[str, Any] | None]:
        signature = definition.signature or {}
        expected_range = definition.range_ref
        expected_headers = signature.get("headers", [])
        candidate_ranges = {candidate["rangeRef"] for candidate in detected_candidates}
        candidate_ranges.add(expected_range)
        scored: list[tuple[float, str, list[str], list[str]]] = []
        for candidate_range in candidate_ranges:
            headers = self._headers_for_candidate(ws, candidate_range, signature)
            score, reasons = self._table_match_score(
                expected_range, candidate_range, expected_headers, headers, signature
            )
            scored.append((score, candidate_range, headers, reasons))
        scored.sort(reverse=True, key=lambda item: item[0])
        best = scored[0] if scored else None
        ambiguity_pool = scored
        # The exact expected rectangle can mask an inserted trailing row/column because
        # reading that rectangle still yields valid headers. Prefer a detected block with
        # the same anchor when its dimensions changed, and treat both as one hypothesis.
        expected_min_col, expected_min_row, expected_max_col, expected_max_row = range_boundaries(
            expected_range
        )
        anchored_structural = []
        for item in scored:
            if item[1] == expected_range or item[0] < 0.62:
                continue
            min_col, min_row, max_col, max_row = range_boundaries(item[1])
            if (min_col, min_row) == (expected_min_col, expected_min_row) and (
                max_col != expected_max_col or max_row != expected_max_row
            ):
                anchored_structural.append(item)
        if anchored_structural:
            best = max(anchored_structural, key=lambda item: item[0])
            ambiguity_pool = [item for item in scored if item[1] != expected_range]
        if best is None or best[0] < 0.48:
            anomaly = self._anomaly(
                "TABLE_MISSING",
                "BLOCKING" if definition.required else "WARNING",
                f"Le tableau « {definition.name} » est introuvable.",
                sheet_name=ws.title,
                table=definition,
                expected=expected_headers,
                actual=None,
                expected_coordinates=expected_range,
                suggestion="Restaurer le tableau ou sélectionner manuellement sa nouvelle zone.",
                candidates=self._candidate_details(scored[:5]),
                expected_preview=signature.get("preview"),
            )
            return [anomaly], None
        if (
            len(ambiguity_pool) > 1
            and ambiguity_pool[1][0] >= 0.58
            and ambiguity_pool[0][0] - ambiguity_pool[1][0] <= 0.06
            and ambiguity_pool[0][1] != ambiguity_pool[1][1]
        ):
            best = ambiguity_pool[0]
            second = ambiguity_pool[1]
            anomaly = self._anomaly(
                "AMBIGUOUS_TABLE_MATCH",
                "BLOCKING",
                f"Plusieurs zones ressemblent au tableau « {definition.name} ».",
                sheet_name=ws.title,
                table=definition,
                expected=expected_headers,
                actual=[best[2], second[2]],
                expected_coordinates=expected_range,
                suggestion="Choisir manuellement la zone correcte avant extraction.",
                confidence=best[0],
                match_reasons=best[3],
                candidates=self._candidate_details(scored[:5]),
                expected_preview=signature.get("preview"),
            )
            return [anomaly], None

        confidence, actual_range, actual_headers, reasons = best
        anomalies: list[dict[str, Any]] = []
        if actual_range != expected_range:
            anomalies.append(
                self._anomaly(
                    "TABLE_MOVED",
                    "WARNING",
                    f"Le tableau « {definition.name} » a été retrouvé à une autre position.",
                    sheet_name=ws.title,
                    table=definition,
                    expected=expected_range,
                    actual=actual_range,
                    expected_coordinates=expected_range,
                    actual_coordinates=actual_range,
                    suggestion="Confirmer le déplacement ou replacer le tableau.",
                    confidence=confidence,
                    match_reasons=reasons,
                    expected_preview=signature.get("preview"),
                    actual_preview={"headers": actual_headers, "rows": self._preview(ws, actual_range)},
                )
            )

        anomalies.extend(
            self._compare_columns(ws, definition, actual_range, expected_headers, actual_headers)
        )
        anomalies.extend(self._compare_rows(ws, definition, actual_range))
        anomalies.extend(self._compare_formulas(ws, definition, actual_range))
        anomalies.extend(self._compare_required_cells(ws, definition, actual_range))
        anomalies.extend(self._compare_table_merges(ws, definition, actual_range))

        if any(item["severity"] == "BLOCKING" for item in anomalies):
            return anomalies, None
        extraction = self._extract_table(ws, definition, actual_range, actual_headers, anomalies)
        return anomalies, extraction

    def _compare_columns(
        self,
        ws: Any,
        definition: TableDefinition,
        actual_range: str,
        expected_headers: list[str],
        actual_headers: list[str],
    ) -> list[dict[str, Any]]:
        expected_normalized = [normalize_text(value) for value in expected_headers]
        actual_normalized = [normalize_text(value) for value in actual_headers]
        expected_counter = Counter(expected_normalized)
        actual_counter = Counter(actual_normalized)
        missing = list((expected_counter - actual_counter).elements())
        added = list((actual_counter - expected_counter).elements())
        anomalies: list[dict[str, Any]] = []
        severity = "ERROR" if definition.structure_mode == StructureMode.STRICT.value else "WARNING"
        if missing:
            anomalies.append(
                self._table_diff(
                    "COLUMN_REMOVED", severity, definition, ws.title, actual_range, missing, None
                )
            )
        if added:
            anomalies.append(
                self._table_diff(
                    "COLUMN_ADDED", severity, definition, ws.title, actual_range, None, added
                )
            )
        if not missing and not added and expected_normalized != actual_normalized:
            anomalies.append(
                self._table_diff(
                    "COLUMN_ORDER_CHANGED",
                    severity,
                    definition,
                    ws.title,
                    actual_range,
                    expected_headers,
                    actual_headers,
                )
            )
        if missing and added and len(missing) == len(added):
            rename_pairs = []
            remaining = added.copy()
            for old in missing:
                if not remaining:
                    break
                new = max(remaining, key=lambda value: similarity(old, value))
                confidence = similarity(old, new)
                if confidence >= 0.58:
                    rename_pairs.append({"expected": old, "actual": new, "confidence": confidence})
                    remaining.remove(new)
            if rename_pairs:
                anomalies.append(
                    self._table_diff(
                        "COLUMN_RENAMED",
                        severity,
                        definition,
                        ws.title,
                        actual_range,
                        missing,
                        rename_pairs,
                    )
                )
        if expected_normalized != actual_normalized:
            anomalies.append(
                self._table_diff(
                    "HEADER_CHANGED",
                    severity,
                    definition,
                    ws.title,
                    actual_range,
                    expected_headers,
                    actual_headers,
                )
            )
        return anomalies

    def _compare_rows(
        self, ws: Any, definition: TableDefinition, actual_range: str
    ) -> list[dict[str, Any]]:
        signature = definition.signature or {}
        expected_keys = [
            tuple(normalize_text(value) for value in item.get("values", []))
            for item in signature.get("rowKeys", [])
        ]
        if not expected_keys or not definition.key_columns:
            return []
        expected_min_col, expected_min_row, _, _ = range_boundaries(definition.range_ref)
        actual_min_col, actual_min_row, actual_max_col, actual_max_row = range_boundaries(actual_range)
        relative_key_columns = [
            column_index_from_string(column) - expected_min_col
            for column in definition.key_columns
        ]
        actual_data_start = actual_min_row + (definition.data_start_row - expected_min_row)
        total_offsets = {row - expected_min_row for row in definition.total_rows}
        actual_keys: list[tuple[str, ...]] = []
        for row_number in range(actual_data_start, actual_max_row + 1):
            if row_number - actual_min_row in total_offsets:
                continue
            values = []
            for offset in relative_key_columns:
                column = actual_min_col + offset
                values.append(
                    normalize_text(ws.cell(row_number, column).value)
                    if column <= actual_max_col
                    else ""
                )
            key = tuple(values)
            if any(key):
                actual_keys.append(key)

        expected_counter = Counter(expected_keys)
        actual_counter = Counter(actual_keys)
        removed = [list(item) for item in (expected_counter - actual_counter).elements()]
        added = [list(item) for item in (actual_counter - expected_counter).elements()]
        if definition.structure_mode == StructureMode.SEMI_DYNAMIC.value:
            return []
        anomalies: list[dict[str, Any]] = []
        if removed:
            anomalies.append(
                self._table_diff(
                    "ROW_REMOVED", "ERROR", definition, ws.title, actual_range, removed, None
                )
            )
        if added:
            anomalies.append(
                self._table_diff(
                    "ROW_ADDED", "ERROR", definition, ws.title, actual_range, None, added
                )
            )
        if not removed and not added and expected_keys != actual_keys:
            anomalies.append(
                self._table_diff(
                    "ROW_ORDER_CHANGED",
                    "ERROR",
                    definition,
                    ws.title,
                    actual_range,
                    [list(item) for item in expected_keys],
                    [list(item) for item in actual_keys],
                )
            )
        return anomalies

    def _compare_formulas(
        self, ws: Any, definition: TableDefinition, actual_range: str
    ) -> list[dict[str, Any]]:
        signature = definition.signature or {}
        formulas = signature.get("requiredFormulas", [])
        if not formulas and definition.structure_mode == StructureMode.STRICT.value:
            formulas = signature.get("formulas", [])
        actual_min_col, actual_min_row, _, _ = range_boundaries(actual_range)
        anomalies: list[dict[str, Any]] = []
        for item in formulas:
            row = actual_min_row + int(item.get("relativeRow", 0))
            column = actual_min_col + int(item.get("relativeColumn", 0))
            cell = ws.cell(row=row, column=column)
            expected_formula = str(item.get("formula") or "")
            actual_formula = cell.value if isinstance(cell.value, str) else ""
            if not actual_formula.startswith("="):
                anomalies.append(
                    self._table_diff(
                        "FORMULA_MISSING",
                        "ERROR",
                        definition,
                        ws.title,
                        actual_range,
                        expected_formula,
                        json_value(cell.value),
                        actual_coordinate=cell.coordinate,
                    )
                )
            elif normalize_text(expected_formula) != normalize_text(actual_formula):
                anomalies.append(
                    self._table_diff(
                        "FORMULA_CHANGED",
                        "ERROR",
                        definition,
                        ws.title,
                        actual_range,
                        expected_formula,
                        actual_formula,
                        actual_coordinate=cell.coordinate,
                    )
                )
        return anomalies

    def _compare_required_cells(
        self, ws: Any, definition: TableDefinition, actual_range: str
    ) -> list[dict[str, Any]]:
        signature = definition.signature or {}
        expected_min_col, expected_min_row, _, _ = range_boundaries(definition.range_ref)
        actual_min_col, actual_min_row, _, _ = range_boundaries(actual_range)
        anomalies: list[dict[str, Any]] = []
        for item in signature.get("requiredCells", []):
            coordinate = item.get("coordinate")
            if not coordinate:
                continue
            original = ws[coordinate]
            row = actual_min_row + original.row - expected_min_row
            column = actual_min_col + original.column - expected_min_col
            actual_cell = ws.cell(row, column)
            expected = item.get("expected")
            actual = json_value(actual_cell.value)
            if normalize_text(expected) != normalize_text(actual):
                anomalies.append(
                    self._table_diff(
                        "KEY_CELL_CHANGED",
                        "ERROR",
                        definition,
                        ws.title,
                        actual_range,
                        expected,
                        actual,
                        actual_coordinate=actual_cell.coordinate,
                    )
                )
        return anomalies

    def _compare_table_merges(
        self, ws: Any, definition: TableDefinition, actual_range: str
    ) -> list[dict[str, Any]]:
        signature = definition.signature or {}
        expected_min_col, expected_min_row, _, _ = range_boundaries(definition.range_ref)
        actual_min_col, actual_min_row, actual_max_col, actual_max_row = range_boundaries(actual_range)
        expected_relative = set()
        for value in signature.get("mergedRanges", []):
            min_col, min_row, max_col, max_row = range_boundaries(value)
            expected_relative.add(
                (
                    min_row - expected_min_row,
                    min_col - expected_min_col,
                    max_row - expected_min_row,
                    max_col - expected_min_col,
                )
            )
        actual_relative = set()
        for merged in ws.merged_cells.ranges:
            if (
                merged.max_row < actual_min_row
                or merged.min_row > actual_max_row
                or merged.max_col < actual_min_col
                or merged.min_col > actual_max_col
            ):
                continue
            actual_relative.add(
                (
                    merged.min_row - actual_min_row,
                    merged.min_col - actual_min_col,
                    merged.max_row - actual_min_row,
                    merged.max_col - actual_min_col,
                )
            )
        if expected_relative == actual_relative:
            return []
        return [
            self._table_diff(
                "MERGED_CELL_CHANGED",
                "WARNING",
                definition,
                ws.title,
                actual_range,
                sorted(expected_relative),
                sorted(actual_relative),
            )
        ]

    def _extract_table(
        self,
        ws: Any,
        definition: TableDefinition,
        actual_range: str,
        headers: list[str],
        anomalies: list[dict[str, Any]],
    ) -> dict[str, Any]:
        expected_min_col, expected_min_row, _, expected_max_row = range_boundaries(
            definition.range_ref
        )
        min_col, min_row, max_col, max_row = range_boundaries(actual_range)
        data_start = min_row + definition.data_start_row - expected_min_row
        if definition.structure_mode == StructureMode.STRICT.value and definition.data_end_row:
            data_end = min(max_row, min_row + definition.data_end_row - expected_min_row)
        else:
            bottom_reserved = max(0, expected_max_row - (definition.data_end_row or expected_max_row))
            data_end = max(data_start - 1, max_row - bottom_reserved)
        total_offsets = {row - expected_min_row for row in definition.total_rows}
        rows: list[dict[str, Any]] = []
        coordinates: list[list[str]] = []
        formulas: list[dict[str, Any]] = []
        for row_number in range(data_start, data_end + 1):
            if row_number - min_row in total_offsets:
                continue
            values: list[Any] = []
            row_coordinates: list[str] = []
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row_number, col)
                values.append(json_value(cell.value))
                row_coordinates.append(cell.coordinate)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append({"coordinate": cell.coordinate, "formula": cell.value})
            if any(value not in (None, "") for value in values):
                usable_headers = headers[: len(values)]
                if len(usable_headers) < len(values):
                    usable_headers += [
                        get_column_letter(min_col + index)
                        for index in range(len(usable_headers), len(values))
                    ]
                rows.append(dict(zip(usable_headers, values, strict=True)))
                coordinates.append(row_coordinates)
        warnings = [
            item["description"]
            for item in anomalies
            if item["severity"] in {"WARNING", "INFO"}
        ]
        return {
            "tableDefinitionId": definition.id,
            "sheetName": ws.title,
            "tableName": definition.name,
            "sourceRange": actual_range,
            "headers": headers,
            "rows": rows,
            "cellCoordinates": coordinates,
            "formulas": formulas,
            "warnings": warnings,
        }

    def _headers_for_candidate(
        self, ws: Any, candidate_range: str, signature: dict[str, Any]
    ) -> list[str]:
        min_col, min_row, max_col, _ = range_boundaries(candidate_range)
        original_min_row = int(signature.get("minRow", min_row))
        header_offsets = [
            int(row) - original_min_row for row in signature.get("headerRows", [original_min_row])
        ]
        result: list[str] = []
        for col in range(min_col, max_col + 1):
            parts = []
            for offset in header_offsets:
                value = ws.cell(min_row + offset, col).value
                if value not in (None, ""):
                    parts.append(str(value).strip())
            result.append(" / ".join(parts) or get_column_letter(col))
        return result

    @staticmethod
    def _table_match_score(
        expected_range: str,
        candidate_range: str,
        expected_headers: list[str],
        actual_headers: list[str],
        signature: dict[str, Any],
    ) -> tuple[float, list[str]]:
        expected = [normalize_text(value) for value in expected_headers]
        actual = [normalize_text(value) for value in actual_headers]
        set_score = len(set(expected) & set(actual)) / max(1, len(set(expected) | set(actual)))
        sequence_score = SequenceMatcher(None, expected, actual).ratio()
        header_score = max(set_score, sequence_score)
        expected_min_col, expected_min_row, expected_max_col, expected_max_row = range_boundaries(
            expected_range
        )
        actual_min_col, actual_min_row, actual_max_col, actual_max_row = range_boundaries(
            candidate_range
        )
        expected_width = expected_max_col - expected_min_col + 1
        actual_width = actual_max_col - actual_min_col + 1
        expected_height = expected_max_row - expected_min_row + 1
        actual_height = actual_max_row - actual_min_row + 1
        width_score = min(expected_width, actual_width) / max(expected_width, actual_width)
        height_score = min(expected_height, actual_height) / max(expected_height, actual_height)
        distance = abs(expected_min_col - actual_min_col) + abs(expected_min_row - actual_min_row)
        proximity = 1 / (1 + distance / 10)
        score = header_score * 0.68 + width_score * 0.14 + height_score * 0.08 + proximity * 0.10
        reasons = [f"Similarité des en-têtes : {header_score:.0%}"]
        if width_score >= 0.9:
            reasons.append("Nombre de colonnes cohérent")
        if height_score >= 0.8:
            reasons.append("Hauteur de zone cohérente")
        if distance == 0:
            reasons.append("Position attendue")
        elif proximity >= 0.5:
            reasons.append("Proximité avec la position initiale")
        return round(score, 3), reasons

    @staticmethod
    def _candidate_details(
        scored: list[tuple[float, str, list[str], list[str]]]
    ) -> list[dict[str, Any]]:
        return [
            {
                "rangeRef": range_ref,
                "confidence": score,
                "headers": headers,
                "reasons": reasons,
            }
            for score, range_ref, headers, reasons in scored
        ]

    @staticmethod
    def _match_renamed_sheets(
        missing: list[SheetDefinition],
        added: list[str],
        actual_metadata: dict[str, dict[str, Any]],
    ) -> list[tuple[SheetDefinition, str, float, list[str]]]:
        options: list[tuple[float, SheetDefinition, str, list[str]]] = []
        for sheet in missing:
            expected_labels = set((sheet.structural_signature or {}).get("labels", []))
            for actual_name in added:
                actual = actual_metadata.get(actual_name, {})
                actual_labels = set(
                    (actual.get("structuralSignature") or {}).get("labels", [])
                )
                name_score = similarity(sheet.name, actual_name)
                actual_index = int(actual.get("originalIndex", 0))
                position_score = 1 / (1 + abs(sheet.original_index - actual_index))
                label_score = len(expected_labels & actual_labels) / max(
                    1, len(expected_labels | actual_labels)
                )
                score = name_score * 0.5 + position_score * 0.2 + label_score * 0.3
                reasons = [f"Similarité du nom : {name_score:.0%}"]
                if position_score >= 0.5:
                    reasons.append("Position de feuille proche")
                if label_score > 0:
                    reasons.append(f"Intitulés communs : {label_score:.0%}")
                options.append((score, sheet, actual_name, reasons))
        options.sort(reverse=True, key=lambda item: item[0])
        selected: list[tuple[SheetDefinition, str, float, list[str]]] = []
        used_expected: set[str] = set()
        used_actual: set[str] = set()
        for score, sheet, actual_name, reasons in options:
            if score < 0.56 or sheet.name in used_expected or actual_name in used_actual:
                continue
            selected.append((sheet, actual_name, round(score, 3), reasons))
            used_expected.add(sheet.name)
            used_actual.add(actual_name)
        return selected

    @staticmethod
    def _major_dimension_change(sheet: SheetDefinition, actual_ws: Any) -> bool:
        row_delta = abs(sheet.max_row - actual_ws.max_row)
        col_delta = abs(sheet.max_column - actual_ws.max_column)
        row_ratio = row_delta / max(1, sheet.max_row)
        col_ratio = col_delta / max(1, sheet.max_column)
        return (row_delta >= 5 and row_ratio >= 0.25) or (
            col_delta >= 3 and col_ratio >= 0.25
        )

    @staticmethod
    def _named_range_map(items: list[dict[str, Any]]) -> dict[str, Any]:
        return {
            item.get("name", ""): item.get("destinations", [])
            for item in items
            if item.get("name")
        }

    @staticmethod
    def _preview(ws: Any, range_ref: str) -> list[list[Any]]:
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
        return [
            [json_value(ws.cell(row, col).value) for col in range(min_col, max_col + 1)]
            for row in range(min_row, min(max_row, min_row + 5) + 1)
        ]

    @staticmethod
    def _anomaly(
        category: str,
        severity: str,
        description: str,
        *,
        sheet_name: str | None = None,
        table: TableDefinition | None = None,
        expected: Any = None,
        actual: Any = None,
        expected_coordinates: str | None = None,
        actual_coordinates: str | None = None,
        suggestion: str | None = None,
        confidence: float | None = None,
        match_reasons: list[str] | None = None,
        candidates: list[dict[str, Any]] | None = None,
        expected_preview: dict[str, Any] | None = None,
        actual_preview: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        return {
            "sheetName": sheet_name,
            "tableDefinitionId": table.id if table else None,
            "tableName": table.name if table else None,
            "category": category,
            "severity": severity,
            "description": description,
            "expected": expected,
            "actual": actual,
            "expectedCoordinates": expected_coordinates,
            "actualCoordinates": actual_coordinates,
            "suggestion": suggestion,
            "confidence": confidence,
            "matchReasons": match_reasons or [],
            "candidates": candidates or [],
            "expectedPreview": expected_preview,
            "actualPreview": actual_preview,
        }

    def _table_diff(
        self,
        category: str,
        severity: str,
        definition: TableDefinition,
        sheet_name: str,
        actual_range: str,
        expected: Any,
        actual: Any,
        actual_coordinate: str | None = None,
    ) -> dict[str, Any]:
        labels = {
            "COLUMN_REMOVED": "Une ou plusieurs colonnes ont été supprimées.",
            "COLUMN_ADDED": "Une ou plusieurs colonnes ont été ajoutées.",
            "COLUMN_ORDER_CHANGED": "L’ordre des colonnes a changé.",
            "COLUMN_RENAMED": "Une ou plusieurs colonnes semblent avoir été renommées.",
            "HEADER_CHANGED": "Les en-têtes du tableau ont changé.",
            "ROW_REMOVED": "Une ou plusieurs lignes structurelles ont été supprimées.",
            "ROW_ADDED": "Une ou plusieurs lignes structurelles ont été ajoutées.",
            "ROW_ORDER_CHANGED": "L’ordre des lignes structurelles a changé.",
            "FORMULA_MISSING": "Une formule obligatoire est absente.",
            "FORMULA_CHANGED": "Une formule obligatoire a changé.",
            "KEY_CELL_CHANGED": "Une cellule structurelle obligatoire a changé.",
            "MERGED_CELL_CHANGED": "Les cellules fusionnées du tableau ont changé.",
        }
        return self._anomaly(
            category,
            severity,
            f"{labels.get(category, 'La structure du tableau a changé')} Tableau « {definition.name} ».",
            sheet_name=sheet_name,
            table=definition,
            expected=expected,
            actual=actual,
            expected_coordinates=definition.range_ref,
            actual_coordinates=actual_coordinate or actual_range,
            suggestion="Comparer la structure détectée avec le mapping validé.",
            expected_preview=(definition.signature or {}).get("preview"),
            actual_preview={"headers": self._headers_for_candidate_placeholder(actual)},
        )

    @staticmethod
    def _headers_for_candidate_placeholder(actual: Any) -> list[Any]:
        return actual if isinstance(actual, list) else []
