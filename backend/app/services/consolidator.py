from __future__ import annotations

import copy
import hashlib
import io
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell

from ..errors import DomainError
from .excel_reader import safe_load_workbook


INVALID_SHEET_CHARACTERS = re.compile(r"[\\/*?:\[\]]")


def _ascii_token(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    normalized = "".join(char for char in normalized if not unicodedata.combining(char))
    normalized = INVALID_SHEET_CHARACTERS.sub("_", normalized)
    normalized = re.sub(r"\s+", "_", normalized.strip())
    normalized = re.sub(r"_+", "_", normalized)
    normalized = normalized.strip("_' ")
    return normalized or "SHEET"


def safe_sheet_name(original: str, country_suffix: str, used: set[str]) -> str:
    base = _ascii_token(original)
    suffix = _ascii_token(country_suffix).upper()
    suffix_part = f"_{suffix}"
    candidate = f"{base[: max(1, 31 - len(suffix_part))]}{suffix_part}"[:31]
    candidate = candidate.rstrip("'") or "SHEET"
    lowered = {value.casefold() for value in used}
    if candidate.casefold() not in lowered:
        used.add(candidate)
        return candidate
    counter = 2
    while True:
        collision_suffix = f"_{counter}"
        attempt = f"{candidate[: 31 - len(collision_suffix)]}{collision_suffix}"
        if attempt.casefold() not in lowered:
            used.add(attempt)
            return attempt
        counter += 1


@dataclass(frozen=True)
class ConsolidationSource:
    country_id: str
    country_name: str
    country_code: str | None
    file_version_id: str
    original_filename: str
    path: Path


class OpenpyxlWorkbookConsolidator:
    """Copies worksheet-level content while reporting OOXML features not copied."""

    def consolidate(
        self, sources: list[ConsolidationSource]
    ) -> tuple[bytes, dict[str, Any], str]:
        if not sources:
            raise DomainError(
                "NO_ELIGIBLE_FILES",
                "No eligible country workbook was selected for consolidation",
                status_code=422,
            )
        destination = Workbook()
        placeholder = destination.active
        placeholder.title = "__placeholder__"
        used_names: set[str] = set()
        globally_used_table_names: set[str] = set()
        report: dict[str, Any] = {
            "countriesIncluded": [],
            "filesUsed": [],
            "sheetsCopied": [],
            "sheetsIgnored": [],
            "nameMappings": [],
            "warnings": [],
            "errors": [],
            "fidelity": {
                "values": True,
                "formulas": True,
                "styles": True,
                "mergedCells": True,
                "comments": True,
                "hyperlinks": True,
                "charts": False,
                "images": False,
                "embeddedObjects": False,
                "externalConnections": False,
                "macros": False,
            },
        }
        try:
            for source in sources:
                workbook = safe_load_workbook(source.path)
                try:
                    suffix = source.country_code or source.country_name
                    mapping = {
                        sheet_name: safe_sheet_name(sheet_name, suffix, used_names)
                        for sheet_name in workbook.sheetnames
                    }
                    report["countriesIncluded"].append(
                        {
                            "id": source.country_id,
                            "name": source.country_name,
                            "code": source.country_code,
                        }
                    )
                    report["filesUsed"].append(
                        {
                            "fileVersionId": source.file_version_id,
                            "originalFilename": source.original_filename,
                        }
                    )
                    if list(workbook.defined_names.values()):
                        report["warnings"].append(
                            {
                                "country": source.country_name,
                                "code": "DEFINED_NAMES_NOT_COPIED",
                                "message": "Les plages nommées du classeur source ne sont pas recopiées.",
                            }
                        )
                    for source_sheet in workbook.worksheets:
                        target_name = mapping[source_sheet.title]
                        target_sheet = destination.create_sheet(target_name)
                        sheet_warnings = self._copy_sheet(
                            source_sheet,
                            target_sheet,
                            mapping,
                            globally_used_table_names,
                        )
                        report["nameMappings"].append(
                            {
                                "country": source.country_name,
                                "originalSheetName": source_sheet.title,
                                "consolidatedSheetName": target_name,
                            }
                        )
                        report["sheetsCopied"].append(
                            {
                                "country": source.country_name,
                                "original": source_sheet.title,
                                "consolidated": target_name,
                            }
                        )
                        report["warnings"].extend(
                            {
                                "country": source.country_name,
                                "sheet": source_sheet.title,
                                **warning,
                            }
                            for warning in sheet_warnings
                        )
                except Exception as exc:
                    report["errors"].append(
                        {
                            "country": source.country_name,
                            "fileVersionId": source.file_version_id,
                            "message": str(exc),
                        }
                    )
                    raise DomainError(
                        "CONSOLIDATION_COPY_FAILED",
                        f"Copy failed for country {source.country_name}",
                        status_code=422,
                        details={"countryId": source.country_id},
                    ) from exc
                finally:
                    workbook.close()

            destination.remove(placeholder)
            if not destination.worksheets:
                raise DomainError("EMPTY_CONSOLIDATION", "No worksheet could be copied")
            if all(sheet.sheet_state != "visible" for sheet in destination.worksheets):
                destination.worksheets[0].sheet_state = "visible"
                report["warnings"].append(
                    {
                        "code": "FIRST_SHEET_FORCED_VISIBLE",
                        "message": "Excel exige au moins une feuille visible; la première a été rendue visible.",
                    }
                )
            destination.properties.creator = "POPS Workbook Platform"
            destination.properties.title = "POPS consolidated workbook"
            buffer = io.BytesIO()
            destination.save(buffer)
            payload = buffer.getvalue()
            report["output"] = {
                "sheetCount": len(destination.worksheets),
                "sizeBytes": len(payload),
                "sha256": hashlib.sha256(payload).hexdigest(),
            }
            filename = "POPS_consolidated.xlsx"
            return payload, report, filename
        finally:
            destination.close()

    def _copy_sheet(
        self,
        source: Any,
        target: Any,
        sheet_mapping: dict[str, str],
        globally_used_table_names: set[str],
    ) -> list[dict[str, str]]:
        warnings: list[dict[str, str]] = []
        target.sheet_state = source.sheet_state
        for row in source.iter_rows():
            for source_cell in row:
                if isinstance(source_cell, MergedCell):
                    continue
                target_cell = target.cell(source_cell.row, source_cell.column)
                value = source_cell.value
                if isinstance(value, str) and value.startswith("="):
                    value = self._rewrite_formula(value, sheet_mapping)
                target_cell.value = value
                if source_cell.has_style:
                    target_cell.font = copy.copy(source_cell.font)
                    target_cell.fill = copy.copy(source_cell.fill)
                    target_cell.border = copy.copy(source_cell.border)
                    target_cell.alignment = copy.copy(source_cell.alignment)
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = copy.copy(source_cell.protection)
                if source_cell.comment is not None:
                    target_cell.comment = copy.copy(source_cell.comment)
                if source_cell.hyperlink is not None:
                    target_cell._hyperlink = copy.copy(source_cell.hyperlink)

        for key, dimension in source.column_dimensions.items():
            target_dimension = target.column_dimensions[key]
            for attribute in ("width", "hidden", "bestFit", "outlineLevel", "collapsed"):
                setattr(target_dimension, attribute, getattr(dimension, attribute))
        for index, dimension in source.row_dimensions.items():
            target_dimension = target.row_dimensions[index]
            for attribute in ("height", "hidden", "outlineLevel", "collapsed"):
                setattr(target_dimension, attribute, getattr(dimension, attribute))
        for merged in source.merged_cells.ranges:
            target.merge_cells(str(merged))

        target.freeze_panes = source.freeze_panes
        target.auto_filter.ref = source.auto_filter.ref
        target.sheet_format = copy.copy(source.sheet_format)
        target.sheet_properties = copy.copy(source.sheet_properties)
        target.page_margins = copy.copy(source.page_margins)
        target.page_setup = copy.copy(source.page_setup)
        target.print_options = copy.copy(source.print_options)
        target.sheet_view.showGridLines = source.sheet_view.showGridLines
        target.sheet_view.zoomScale = source.sheet_view.zoomScale
        if source.print_area:
            target.print_area = source.print_area
        if source.print_title_rows:
            target.print_title_rows = source.print_title_rows
        if source.print_title_cols:
            target.print_title_cols = source.print_title_cols

        for validation in source.data_validations.dataValidation:
            target.add_data_validation(copy.copy(validation))
        try:
            for conditional_format in source.conditional_formatting:
                rules = source.conditional_formatting[conditional_format]
                for rule in rules:
                    target.conditional_formatting.add(
                        str(conditional_format.sqref), copy.copy(rule)
                    )
        except Exception:
            warnings.append(
                {
                    "code": "CONDITIONAL_FORMATTING_PARTIAL",
                    "message": "Une mise en forme conditionnelle n’a pas pu être recopiée.",
                }
            )

        for table_name in source.tables:
            try:
                table_copy = copy.deepcopy(source.tables[table_name])
                base_name = re.sub(r"[^A-Za-z0-9_]", "_", table_copy.displayName)
                candidate = base_name
                suffix = 2
                while candidate.casefold() in globally_used_table_names:
                    candidate = f"{base_name}_{suffix}"
                    suffix += 1
                table_copy.name = candidate
                table_copy.displayName = candidate
                globally_used_table_names.add(candidate.casefold())
                target.add_table(table_copy)
            except Exception:
                warnings.append(
                    {
                        "code": "NATIVE_TABLE_NOT_COPIED",
                        "message": f"Le tableau Excel natif « {table_name} » n’a pas pu être recopié.",
                    }
                )
        if getattr(source, "_charts", []):
            warnings.append(
                {
                    "code": "CHARTS_NOT_COPIED",
                    "message": f"{len(source._charts)} graphique(s) non recopié(s).",
                }
            )
        if getattr(source, "_images", []):
            warnings.append(
                {
                    "code": "IMAGES_NOT_COPIED",
                    "message": f"{len(source._images)} image(s) non recopiée(s).",
                }
            )
        return warnings

    @staticmethod
    def _rewrite_formula(formula: str, mapping: dict[str, str]) -> str:
        result = formula
        for old_name in sorted(mapping, key=len, reverse=True):
            new_name = mapping[old_name].replace("'", "''")
            old_quoted = old_name.replace("'", "''")
            result = result.replace(f"'{old_quoted}'!", f"'{new_name}'!")
            result = re.sub(
                rf"(?<![A-Za-z0-9_'\]]){re.escape(old_name)}!",
                f"'{new_name}'!",
                result,
            )
        return result
