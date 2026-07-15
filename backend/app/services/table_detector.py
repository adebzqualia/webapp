from __future__ import annotations

import hashlib
from collections import deque
from pathlib import Path
from typing import Any

from openpyxl.utils import get_column_letter, range_boundaries

from ..errors import DomainError
from .excel_reader import cell_style, json_value, normalize_text, safe_load_workbook


class HeuristicTableDetector:
    """Suggests table-like rectangles; it never persists or validates a candidate."""

    def __init__(self, max_scan_cells: int = 500_000) -> None:
        self.max_scan_cells = max_scan_cells

    def detect_candidates(
        self,
        file_path: Path,
        sheet_names: list[str] | None = None,
    ) -> list[dict[str, Any]]:
        workbook = safe_load_workbook(file_path)
        try:
            selected = sheet_names or workbook.sheetnames
            candidates: list[dict[str, Any]] = []
            definitions = list(workbook.defined_names.values())
            for sheet_name in selected:
                if sheet_name not in workbook.sheetnames:
                    raise DomainError("SHEET_NOT_FOUND", "Worksheet not found", status_code=404)
                ws = workbook[sheet_name]

                for table_name in ws.tables:
                    table = ws.tables[table_name]
                    candidates.append(
                        self._make_candidate(
                            ws,
                            table.ref.replace("$", ""),
                            0.99,
                            ["Tableau Excel natif déclaré dans le classeur"],
                            "NATIVE_TABLE",
                        )
                    )

                for definition in definitions:
                    try:
                        destinations = list(definition.destinations)
                    except (AttributeError, TypeError, ValueError):
                        continue
                    for destination_sheet, ref in destinations:
                        if destination_sheet == sheet_name and ":" in ref:
                            range_ref = ref.replace("$", "")
                            try:
                                min_col, min_row, max_col, max_row = range_boundaries(range_ref)
                            except ValueError:
                                continue
                            if max_col > min_col and max_row > min_row:
                                candidates.append(
                                    self._make_candidate(
                                        ws,
                                        range_ref,
                                        0.93,
                                        [f"Plage nommée « {definition.name} »"],
                                        "NAMED_RANGE",
                                    )
                                )

                candidates.extend(self._component_candidates(ws))
            return self._deduplicate(candidates)
        finally:
            workbook.close()

    def _component_candidates(self, ws: Any) -> list[dict[str, Any]]:
        if ws.max_row * ws.max_column > self.max_scan_cells:
            # Native tables/named ranges are still returned by the caller.
            return []
        occupied: set[tuple[int, int]] = set()
        for row in ws.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    occupied.add((cell.row, cell.column))

        result: list[dict[str, Any]] = []
        while occupied:
            seed = occupied.pop()
            queue: deque[tuple[int, int]] = deque([seed])
            component = [seed]
            while queue:
                row, col = queue.popleft()
                for neighbor in (
                    (row - 1, col),
                    (row + 1, col),
                    (row, col - 1),
                    (row, col + 1),
                ):
                    if neighbor in occupied:
                        occupied.remove(neighbor)
                        queue.append(neighbor)
                        component.append(neighbor)
            min_row = min(value[0] for value in component)
            max_row = max(value[0] for value in component)
            min_col = min(value[1] for value in component)
            max_col = max(value[1] for value in component)
            height = max_row - min_row + 1
            width = max_col - min_col + 1
            if height < 2 or width < 2 or len(component) < 4:
                continue
            area = height * width
            density = len(component) / area
            if density < 0.18:
                continue
            range_ref = (
                f"{get_column_letter(min_col)}{min_row}:"
                f"{get_column_letter(max_col)}{max_row}"
            )
            score, reasons = self._score_component(
                ws, min_row, max_row, min_col, max_col, density
            )
            result.append(
                self._make_candidate(ws, range_ref, score, reasons, "HEURISTIC")
            )
        return result

    def _score_component(
        self,
        ws: Any,
        min_row: int,
        max_row: int,
        min_col: int,
        max_col: int,
        density: float,
    ) -> tuple[float, list[str]]:
        score = 0.25 + min(0.3, density * 0.3)
        reasons = [f"Zone rectangulaire non vide (densité {density:.0%})"]
        first_row = [ws.cell(min_row, col) for col in range(min_col, max_col + 1)]
        text_headers = sum(
            1
            for cell in first_row
            if isinstance(cell.value, str) and not cell.value.startswith("=")
        )
        if text_headers >= max(2, len(first_row) // 2):
            score += 0.12
            reasons.append("Première ligne majoritairement textuelle")
        if sum(bool(cell.font.bold) for cell in first_row) >= max(1, len(first_row) // 2):
            score += 0.08
            reasons.append("En-têtes mis en gras")
        if sum(cell_style(cell).get("hasBorder", False) for cell in first_row) >= max(
            1, len(first_row) // 2
        ):
            score += 0.06
            reasons.append("Bordures cohérentes sur les en-têtes")
        style_ids = [cell.style_id for cell in first_row if cell.has_style]
        if style_ids and len(set(style_ids)) <= max(2, len(style_ids) // 2):
            score += 0.05
            reasons.append("Styles homogènes dans la zone")
        formula_count = 0
        for row in ws.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
        ):
            formula_count += sum(
                isinstance(cell.value, str) and cell.value.startswith("=") for cell in row
            )
        if formula_count >= 2:
            score += min(0.08, formula_count / max(1, (max_row - min_row) * 20))
            reasons.append("Formules répétées détectées")
        if max_row - min_row >= 2 and max_col - min_col >= 2:
            score += 0.05
        return round(min(score, 0.9), 3), reasons[:5]

    def _make_candidate(
        self,
        ws: Any,
        range_ref: str,
        confidence: float,
        reasons: list[str],
        source: str,
    ) -> dict[str, Any]:
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
        header_row = self._best_header_row(ws, min_row, min(max_row, min_row + 2), min_col, max_col)
        headers = [
            str(ws.cell(header_row, col).value).strip()
            if ws.cell(header_row, col).value not in (None, "")
            else get_column_letter(col)
            for col in range(min_col, max_col + 1)
        ]
        rows: list[list[Any]] = []
        for row_number in range(header_row + 1, min(max_row, header_row + 5) + 1):
            rows.append(
                [json_value(ws.cell(row_number, col).value) for col in range(min_col, max_col + 1)]
            )
        identifier = hashlib.sha256(
            f"{ws.title}:{range_ref}:{source}".encode("utf-8")
        ).hexdigest()[:20]
        return {
            "id": identifier,
            "sheetName": ws.title,
            "rangeRef": range_ref,
            "confidence": confidence,
            "reasons": reasons,
            "source": source,
            "preview": {"headers": headers, "rows": rows},
        }

    @staticmethod
    def _best_header_row(ws: Any, min_row: int, max_row: int, min_col: int, max_col: int) -> int:
        best = min_row
        best_score = -1.0
        for row_number in range(min_row, max_row + 1):
            cells = [ws.cell(row_number, col) for col in range(min_col, max_col + 1)]
            populated = [cell for cell in cells if cell.value not in (None, "")]
            if not populated:
                continue
            text_ratio = sum(isinstance(cell.value, str) for cell in populated) / len(populated)
            bold_ratio = sum(bool(cell.font.bold) for cell in populated) / len(populated)
            distinct = len({normalize_text(cell.value) for cell in populated}) / len(populated)
            score = len(populated) / len(cells) + text_ratio * 0.5 + bold_ratio * 0.3 + distinct * 0.1
            if score > best_score:
                best_score = score
                best = row_number
        return best

    @staticmethod
    def _deduplicate(candidates: list[dict[str, Any]]) -> list[dict[str, Any]]:
        priority = {"NATIVE_TABLE": 3, "NAMED_RANGE": 2, "HEURISTIC": 1}
        selected: dict[tuple[str, str], dict[str, Any]] = {}
        for candidate in candidates:
            key = (candidate["sheetName"], candidate["rangeRef"])
            previous = selected.get(key)
            if previous is None or priority[candidate["source"]] > priority[previous["source"]]:
                selected[key] = candidate
        return sorted(
            selected.values(),
            key=lambda item: (item["sheetName"], -item["confidence"], item["rangeRef"]),
        )
