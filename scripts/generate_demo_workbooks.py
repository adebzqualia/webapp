"""Generate deterministic POPS workbooks for local demos and manual testing.

The generated country files intentionally cover a conforming import and a file
containing several structural anomalies.  The script never reads production
files and only writes below the selected output directory.
"""

from __future__ import annotations

import argparse
import shutil
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADER_FILL = PatternFill("solid", fgColor="173B57")
SECTION_FILL = PatternFill("solid", fgColor="DCEAF3")
WHITE_BOLD = Font(color="FFFFFF", bold=True)


def style_header(sheet, row: int, start_column: int, end_column: int) -> None:
    for column in range(start_column, end_column + 1):
        cell = sheet.cell(row=row, column=column)
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = Alignment(horizontal="center")


def build_template(path: Path) -> None:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview.merge_cells("B2:H3")
    overview["B2"] = "POPS — Planning & Operational Performance Summary"
    overview["B2"].font = Font(size=16, bold=True, color="173B57")
    overview["B5"] = "Country"
    overview["C5"] = "<to be completed>"
    overview["B7"] = "Reporting period"
    overview["C7"] = "FY 2026"
    overview.sheet_view.showGridLines = False

    financial = workbook.create_sheet("Financial KPIs")
    financial.merge_cells("B2:G3")
    financial["B2"] = "Financial performance"
    financial["B2"].font = Font(size=15, bold=True, color="173B57")

    financial.append([])
    financial["B6"] = "Revenue budget"
    financial["B6"].font = Font(bold=True)
    financial["B6"].fill = SECTION_FILL
    headers = ["Business unit", "Metric", "Budget", "Actual", "Variance", "Variance %"]
    for index, value in enumerate(headers, start=2):
        financial.cell(row=7, column=index, value=value)
    style_header(financial, 7, 2, 7)
    rows = [
        ["Retail", "Revenue", 1_200_000, 1_180_000],
        ["Corporate", "Revenue", 900_000, 930_000],
        ["Digital", "Revenue", 620_000, 665_000],
    ]
    for row_index, values in enumerate(rows, start=8):
        for column_index, value in enumerate(values, start=2):
            financial.cell(row=row_index, column=column_index, value=value)
        financial.cell(row=row_index, column=6, value=f"=E{row_index}-D{row_index}")
        financial.cell(row=row_index, column=7, value=f"=IFERROR(F{row_index}/D{row_index},0)")
    financial["B11"] = "TOTAL"
    financial["D11"] = "=SUM(D8:D10)"
    financial["E11"] = "=SUM(E8:E10)"
    financial["F11"] = "=E11-D11"
    financial["G11"] = "=IFERROR(F11/D11,0)"
    for cell in financial[11][1:7]:
        cell.font = Font(bold=True)
    revenue_table = Table(displayName="RevenueBudget", ref="B7:G10")
    revenue_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False
    )
    financial.add_table(revenue_table)

    financial["B15"] = "Cost base"
    financial["B15"].font = Font(bold=True)
    financial["B15"].fill = SECTION_FILL
    cost_headers = ["Cost center", "Owner", "Budget", "Forecast"]
    for index, value in enumerate(cost_headers, start=2):
        financial.cell(row=16, column=index, value=value)
    style_header(financial, 16, 2, 5)
    cost_rows = [
        ["Technology", "Alex Martin", 420_000, 435_000],
        ["Operations", "Samira Diallo", 310_000, 305_000],
        ["Marketing", "Emma Bernard", 265_000, 280_000],
    ]
    for row_index, values in enumerate(cost_rows, start=17):
        for column_index, value in enumerate(values, start=2):
            financial.cell(row=row_index, column=column_index, value=value)
    cost_table = Table(displayName="CostBase", ref="B16:E19")
    cost_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    financial.add_table(cost_table)

    operations = workbook.create_sheet("Operations")
    operations.merge_cells("B2:F3")
    operations["B2"] = "Operational indicators"
    operations["B2"].font = Font(size=15, bold=True, color="173B57")
    operation_headers = ["Site", "Orders", "On time", "Returns", "Service level"]
    for index, value in enumerate(operation_headers, start=2):
        operations.cell(row=6, column=index, value=value)
    style_header(operations, 6, 2, 6)
    operation_rows = [
        ["North", 12_500, 12_100, 210],
        ["South", 10_100, 9_720, 185],
        ["West", 8_900, 8_610, 142],
    ]
    for row_index, values in enumerate(operation_rows, start=7):
        for column_index, value in enumerate(values, start=2):
            operations.cell(row=row_index, column=column_index, value=value)
        operations.cell(row=row_index, column=6, value=f"=IFERROR(D{row_index}/C{row_index},0)")
    operations.auto_filter.ref = "B6:F9"
    operations.freeze_panes = "B7"

    for sheet in workbook.worksheets:
        for column in range(1, sheet.max_column + 1):
            sheet.column_dimensions[sheet.cell(1, column).column_letter].width = 18
    financial.column_dimensions["B"].width = 22
    overview.sheet_state = "visible"
    workbook.save(path)


def build_country_files(template: Path, output: Path) -> None:
    france = output / "pops_france_conforme.xlsx"
    shutil.copyfile(template, france)
    workbook = load_workbook(france, data_only=False)
    workbook["Overview"]["C5"] = "France"
    workbook["Financial KPIs"]["E8"] = 1_205_000
    workbook["Operations"]["C7"] = 12_850
    workbook.save(france)

    germany = output / "pops_germany_anomalies.xlsx"
    shutil.copyfile(template, germany)
    workbook = load_workbook(germany, data_only=False)
    workbook["Overview"]["C5"] = "Germany"
    financial = workbook["Financial KPIs"]
    financial.title = "Finance KPIs"
    financial.insert_rows(7, amount=2)
    financial["F10"] = None
    financial.insert_cols(5)
    financial["E9"] = "Comment"
    financial["E10"] = "Late source"
    operations = workbook["Operations"]
    operations.sheet_state = "hidden"
    workbook.create_sheet("Local notes")["A1"] = "Country-only content"
    workbook.move_sheet(operations, offset=-1)
    workbook.save(germany)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "demo",
        help="Output directory (default: ./demo)",
    )
    args = parser.parse_args()
    output = args.output.resolve()
    output.mkdir(parents=True, exist_ok=True)
    template = output / "template_pops_demo.xlsx"
    build_template(template)
    build_country_files(template, output)
    print(f"Generated demo workbooks in {output}")


if __name__ == "__main__":
    main()
